# gestionale/views.py

# ==============================================================================
# === IMPORTAZIONI ORGANIZZATE                                              ===
# ==============================================================================
# Standard Library
from dataclasses import field
import json
from datetime import date, timedelta,datetime
today = date.today()
from decimal import Decimal

# Django Core
from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Q, Sum, Value, Case, When, F, Count, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.db.models import Sum, Q
from functools import wraps
from django.views import View
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import AccessMixin
from django.apps import apps

# Librerie di terze parti
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from weasyprint import HTML

# Importazioni delle app locali
from .forms import (
    AliquotaIVAForm, AnagraficaForm, CausaleForm, ContoFinanziarioForm, ContoOperativoForm, DiarioAttivitaForm, DipendenteDettaglioForm,
    DocumentoFilterForm, DocumentoRigaForm, DocumentoTestataForm, MezzoAziendaleForm, ModalitaPagamentoForm,
    PagamentoForm, PartitarioFilterForm, PrimaNotaFilterForm, PrimaNotaForm, ScadenzaPersonaleForm,
    ScadenzarioFilterForm, ScadenzaWizardForm,PrimaNotaUpdateForm,PagamentoUpdateForm, TipoScadenzaPersonaleForm, CantiereForm,
    AnagraficaFilterForm, FascicoloCantiereFilterForm
)
from .models import (
    AliquotaIVA, Anagrafica, Cantiere, Causale, ContoFinanziario,
    ContoOperativo, DiarioAttivita, DipendenteDettaglio, DocumentoRiga,
    DocumentoTestata, MezzoAziendale, ModalitaPagamento, PrimaNota, Scadenza, TipoScadenzaPersonale, ScadenzaPersonale
)
from .report_utils import build_filters_string, generate_excel_report, generate_pdf_report
from tenants.models import Company
from .templatetags import currency_filters

from django.utils.timezone import now
from dateutil.relativedelta import relativedelta
from .forms import AnalisiFilterForm

# ==============================================================================
# === MIXINS E FUNZIONI HELPER GLOBALI                                      ===
# ==============================================================================

class TenantRequiredMixin(LoginRequiredMixin, View):
    """Mixin per assicurare che un tenant sia attivo in sessione."""
    def dispatch(self, request, *args, **kwargs):
        if not request.session.get('active_tenant_id'):
            return redirect(reverse('tenant_selection'))
        return super().dispatch(request, *args, **kwargs)


class AdminRequiredMixin(AccessMixin):
    """Verifica che l'utente abbia il ruolo di 'admin' in sessione."""
    def dispatch(self, request, *args, **kwargs):
        if request.session.get('user_company_role') != 'admin':
            messages.error(request, "Accesso negato. È richiesta l'autorizzazione di un amministratore.")
            # Reindirizza a una pagina sicura, come la dashboard
            return redirect(reverse_lazy('dashboard'))
        return super().dispatch(request, *args, **kwargs)

class RoleRequiredMixin(AccessMixin):
    """Mixin per le viste basate su classi che richiede uno o più ruoli specifici."""
    allowed_roles = [] # Deve essere sovrascritto dalle classi figlie

    def dispatch(self, request, *args, **kwargs):
        user_role = request.session.get('user_company_role')
        if user_role not in self.allowed_roles:
            messages.error(request, "Accesso negato. Non hai i permessi necessari.")
            return redirect(reverse_lazy('dashboard')) # O una pagina di errore permessi
        return super().dispatch(request, *args, **kwargs)

def role_required(allowed_roles=None):
    """Decoratore per le viste basate su funzioni che richiede uno o più ruoli specifici."""
    if allowed_roles is None:
        allowed_roles = []

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            user_role = request.session.get('user_company_role')
            if user_role not in allowed_roles:
                messages.error(request, "Accesso negato. Non hai i permessi necessari.")
                return redirect(reverse_lazy('dashboard')) # O una pagina di errore permessi
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

def clear_doc_wizard_session(session):
    """Pulisce i dati del wizard dalla sessione."""
    session.pop('doc_testata_data', None)
    session.pop('doc_righe_data', None)
    session.pop('doc_scadenze_data', None)

def tenant_required(view_func):
    """Decoratore per le viste basate su funzioni che richiede un tenant attivo in sessione."""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('active_tenant_id'):
            messages.error(request, "Seleziona un'azienda per continuare.")
            return redirect(reverse('tenant_selection'))
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# ==============================================================================
# === VISTE PRINCIPALI, ANAGRAFICHE E DOCUMENTI                             ===
# ==============================================================================

class AnagraficaListView(TenantRequiredMixin, RoleRequiredMixin, View):
    """
    Mostra la lista delle anagrafiche, con paginazione e filtri.
    """
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    template_name = 'gestionale/anagrafica_list.html'
    
    def get(self, request, *args, **kwargs):
        # Inizializza il form con i dati da request.GET
        filter_form = AnagraficaFilterForm(request.GET)
        
        anagrafiche_list = Anagrafica.objects.all().order_by('nome_cognome_ragione_sociale')
        
        # Applica i filtri se il form è valido
        if filter_form.is_valid():
            q = filter_form.cleaned_data.get('q')
            tipo = filter_form.cleaned_data.get('tipo')
            attivo_str = filter_form.cleaned_data.get('attivo')

            if q:
                anagrafiche_list = anagrafiche_list.filter(
                    Q(nome_cognome_ragione_sociale__icontains=q) |
                    Q(citta__icontains=q) |
                    Q(p_iva__icontains=q)
                )
            
            if tipo:
                anagrafiche_list = anagrafiche_list.filter(tipo=tipo)
            
            if attivo_str:
                attivo_bool = (attivo_str == 'true')
                anagrafiche_list = anagrafiche_list.filter(attivo=attivo_bool)

        paginator = Paginator(anagrafiche_list, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'anagrafiche': page_obj,
            'page_obj': page_obj,
            'filter_form': filter_form, # Passa il form al template
        }
        return render(request, self.template_name, context)

class AnagraficaCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin', 'contabile']
    # ... (codice invariato) ...
    model = Anagrafica
    form_class = AnagraficaForm
    template_name = 'gestionale/anagrafica_form.html'

    def get_success_url(self):
        if self.object.tipo == Anagrafica.Tipo.DIPENDENTE:
            return reverse('dipendente_dettaglio_create', kwargs={'anagrafica_id': self.object.pk})
        else:
            return reverse('anagrafica_list')

    def form_valid(self, form):
        """
        Ora questo metodo deve solo impostare i campi utente e lasciare
        che il modello si occupi del resto.
        """
        # Otteniamo l'oggetto senza salvarlo
        anagrafica = form.save(commit=False)
        
        # Impostiamo l'utente
        anagrafica.created_by = self.request.user
        anagrafica.updated_by = self.request.user
        
        # Salviamo. Il metodo save() personalizzato del modello Anagrafica
        # verrà chiamato automaticamente e genererà il codice.
        anagrafica.save()
        
        # Assegniamo l'oggetto alla vista per il reindirizzamento
        self.object = anagrafica
        
        return HttpResponseRedirect(self.get_success_url())

class AnagraficaUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin', 'contabile']
    # ... (codice invariato) ...
    model = Anagrafica
    form_class = AnagraficaForm
    template_name = 'gestionale/anagrafica_form.html'
    success_url = reverse_lazy('anagrafica_list')

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

class AnagraficaToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile']
    # ... (codice invariato) ...
    def post(self, request, *args, **kwargs):
        anagrafica = get_object_or_404(Anagrafica, pk=kwargs.get('pk'))
        anagrafica.attivo = not anagrafica.attivo
        anagrafica.updated_by = request.user
        anagrafica.save()
        return redirect('anagrafica_list')

class DipendenteDettaglioCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin', 'contabile']
    # ... (codice invariato) ...
    model = DipendenteDettaglio
    form_class = DipendenteDettaglioForm
    template_name = 'gestionale/dipendente_dettaglio_form.html'
    success_url = reverse_lazy('anagrafica_list')
    def dispatch(self, request, *args, **kwargs):
        self.anagrafica = get_object_or_404(Anagrafica, pk=kwargs['anagrafica_id'], tipo=Anagrafica.Tipo.DIPENDENTE)
        return super().dispatch(request, *args, **kwargs)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anagrafica'] = self.anagrafica
        return context
    def form_valid(self, form):
        dettaglio = form.save(commit=False)
        dettaglio.anagrafica = self.anagrafica
        dettaglio.save()
        self.object = dettaglio
        return HttpResponseRedirect(self.get_success_url())

# ==============================================================================
# === VISTE WIZARD CREAZIONE DOCUMENTO                                      ===
# ==============================================================================

@login_required
@tenant_required
@role_required(allowed_roles=['admin', 'contabile'])
def documento_create_step1_testata(request):
    """
    Step 1 del wizard: inserimento dati della testata.
    """
    if request.method == 'POST':
        form = DocumentoTestataForm(request.POST)
        if form.is_valid():
            request.session['doc_testata_data'] = {
                'tipo_doc': form.cleaned_data['tipo_doc'],
                'anagrafica_id': form.cleaned_data['anagrafica'].pk,
                'data_documento': form.cleaned_data['data_documento'].isoformat(),
                'modalita_pagamento_id': form.cleaned_data['modalita_pagamento'].pk,
                'cantiere_id': form.cleaned_data['cantiere'].pk if form.cleaned_data['cantiere'] else None,
                'note': form.cleaned_data['note'],
                'numero_documento_manuale': form.cleaned_data.get('numero_documento_manuale'),
            }
            request.session.pop('doc_righe_data', None)
            request.session.pop('doc_scadenze_data', None)
            return redirect(reverse('documento_create_step2_righe'))
    else:
        form = DocumentoTestataForm()

    # Definiamo la lista di tipi passivi e la passiamo al template come JSON
    tipi_passivi = [
        DocumentoTestata.TipoDoc.FATTURA_ACQUISTO, 
        DocumentoTestata.TipoDoc.NOTA_CREDITO_ACQUISTO
    ]
    
    context = {
        'form': form,
        'tipi_passivi_json': json.dumps(tipi_passivi) # Converte la lista Python in una stringa JSON
    }
    return render(request, 'gestionale/documento_create_step1.html', context)

@login_required
@tenant_required
@role_required(allowed_roles=['admin', 'contabile'])
def documento_create_step2_righe(request):
    """
    Step 2 del wizard: inserimento delle righe del documento.
    """
    testata_data = request.session.get('doc_testata_data')
    # Se per qualche motivo mancano i dati del passo 1, torna all'inizio.
    if not testata_data:
        return redirect(reverse('documento_create_step1_testata'))

    righe_data = request.session.get('doc_righe_data', [])
    # === NUOVA LOGICA DI ELIMINAZIONE RIGA (GESTITA CON GET) ===
    riga_da_eliminare_idx = request.GET.get('delete_riga')
    if riga_da_eliminare_idx is not None:
        try:
        # Rimuoviamo la riga dalla lista usando il suo indice
            righe_data.pop(int(riga_da_eliminare_idx))
            request.session['doc_righe_data'] = righe_data
            messages.success(request, "Riga eliminata con successo.")
        except (ValueError, IndexError):
            messages.error(request, "Errore durante l'eliminazione della riga.")
        # Reindirizziamo per pulire l'URL dal parametro 'delete_riga'
        return redirect(reverse('documento_create_step2_righe'))


    if request.method == 'POST':
        # Controlliamo se l'utente vuole andare al passo 3
        if 'prosegui_step3' in request.POST:
            if not righe_data:
                # Non si può procedere senza almeno una riga
                messages.error(request, "Devi inserire almeno una riga per proseguire.")
                # Dobbiamo ricreare il form e il contesto anche in questo caso di errore
                form = DocumentoRigaForm()
            else:
                return redirect(reverse('documento_create_step3_scadenze'))
        else:
            # L'utente sta aggiungendo una nuova riga
            form = DocumentoRigaForm(request.POST)
            if form.is_valid():
                nuova_riga = form.cleaned_data
                
                # Calcoliamo i totali della riga
                imponibile_riga = nuova_riga['quantita'] * nuova_riga['prezzo_unitario']
                aliquota = AliquotaIVA.objects.get(pk=nuova_riga['aliquota_iva'].pk)
                iva_riga = imponibile_riga * (aliquota.valore_percentuale / Decimal(100))
                
                # Aggiungiamo i dati alla sessione
                righe_data.append({
                    'descrizione': nuova_riga['descrizione'],
                    'quantita': str(nuova_riga['quantita']),
                    'prezzo_unitario': str(nuova_riga['prezzo_unitario']),
                    'aliquota_iva_id': aliquota.pk,
                    'aliquota_iva_valore': str(aliquota.valore_percentuale),
                    'imponibile_riga': str(imponibile_riga.quantize(Decimal('0.01'))),
                    'iva_riga': str(iva_riga.quantize(Decimal('0.01'))),
                })
                
                request.session['doc_righe_data'] = righe_data
                # Reindirizziamo alla stessa pagina per "pulire" il form POST
                # e mostrare la nuova riga nella tabella. (Pattern Post-Redirect-Get)
                return redirect(reverse('documento_create_step2_righe'))
    else:
        form = DocumentoRigaForm()
        
    # Calcoliamo i totali attuali per il riepilogo
    totale_imponibile = sum(Decimal(r['imponibile_riga']) for r in righe_data)
    totale_iva = sum(Decimal(r['iva_riga']) for r in righe_data)
    
    context = {
        'form': form,
        'testata_data': testata_data,
        'righe_inserite': righe_data,
        'totale_imponibile': totale_imponibile,
        'totale_iva': totale_iva,
        'totale_documento': totale_imponibile + totale_iva
    }
    return render(request, 'gestionale/documento_create_step2.html', context)

@login_required
@tenant_required
@role_required(allowed_roles=['admin', 'contabile'])
def documento_create_step3_scadenze(request):
    """
    Step 3 del wizard: inserimento scadenze e finalizzazione.
    """
    testata_data = request.session.get('doc_testata_data')
    righe_data = request.session.get('doc_righe_data')
    scadenze_data = request.session.get('doc_scadenze_data', [])

    if not testata_data or not righe_data:
        return redirect(reverse('documento_create_step1_testata'))
    
    # === NUOVA LOGICA DI ELIMINAZIONE SCADENZA (GESTITA CON GET) ===
    scadenza_da_eliminare_idx = request.GET.get('delete_scadenza')
    if scadenza_da_eliminare_idx is not None:
        try:
            scadenze_data.pop(int(scadenza_da_eliminare_idx))
            request.session['doc_scadenze_data'] = scadenze_data
            messages.success(request, "Scadenza eliminata con successo.")
        except (ValueError, IndexError):
            messages.error(request, "Errore durante l'eliminazione della scadenza.")
        return redirect(reverse('documento_create_step3_scadenze'))

    # Riconvertiamo le stringhe dalla sessione in Decimal prima di sommarle.
    totale_documento = sum(Decimal(r['imponibile_riga']) + Decimal(r['iva_riga']) for r in righe_data)
    totale_scadenze = sum(Decimal(s['importo_rata']) for s in scadenze_data)
    residuo_da_scadenzare = totale_documento - totale_scadenze

    if request.method == 'POST':
        # Creiamo un'istanza del form per aggiunta scadenza, che potrebbe servirci
        form = ScadenzaWizardForm(request.POST)

        if 'finalizza_documento' in request.POST:
            if residuo_da_scadenzare != 0:
                messages.error(request, "L'importo delle scadenze non corrisponde al totale del documento.")
                # Non facciamo return qui, lasciamo che la vista prosegua e mostri
                # la pagina con il messaggio di errore e il form vuoto pre-compilato.
                # Per farlo, dobbiamo ricreare il form con i valori di default.
                form = ScadenzaWizardForm(initial=get_scadenza_initial_data(testata_data, scadenze_data, residuo_da_scadenzare))

            else:
                # ... LA LOGICA DI SALVATAGGIO FINALE ... (la copio qui sotto per completezza)
                with transaction.atomic():
                    anagrafica = Anagrafica.objects.get(pk=testata_data['anagrafica_id'])
                    modalita_pagamento = ModalitaPagamento.objects.get(pk=testata_data['modalita_pagamento_id'])
                    cantiere = Cantiere.objects.get(pk=testata_data['cantiere_id']) if testata_data['cantiere_id'] else None

                    # === INIZIO LOGICA DI NUMERAZIONE AUTOMATICA ===
                    tipo_doc = testata_data['tipo_doc']
                    numero_doc_finale = "" # Inizializziamo la variabile

                    active_tenant_id = request.session.get('active_tenant_id')
                    tenant_obj = Company.objects.get(pk=active_tenant_id) if active_tenant_id else None

                    # Definiamo i prefissi per i documenti di vendita
                    prefissi_vendita = {
                        DocumentoTestata.TipoDoc.FATTURA_VENDITA: 'FT',
                        DocumentoTestata.TipoDoc.NOTA_CREDITO_VENDITA: 'NC',
                    }

                    if tipo_doc in prefissi_vendita:
                        prefisso = prefissi_vendita[tipo_doc]
                        anno_corrente = date.today().year
                        
                        # Troviamo l'ultimo documento dello stesso tipo e dello stesso anno
                        ultimo_documento = DocumentoTestata.objects.filter(
                            tipo_doc=tipo_doc,
                            data_documento__year=anno_corrente
                        ).order_by('numero_documento').last()

                        if ultimo_documento and ultimo_documento.numero_documento:
                            # Estraiamo il numero progressivo dall'ultimo numero documento
                            try:
                                ultimo_progressivo = int(ultimo_documento.numero_documento.split('-')[-1])
                                nuovo_progressivo = ultimo_progressivo + 1
                            except (IndexError, ValueError):
                                # Fallback se il formato non è quello atteso
                                nuovo_progressivo = 1
                        else:
                            nuovo_progressivo = 1
                        
                        # Componiamo il numero finale
                        numero_doc_finale = f"{prefisso}-{anno_corrente}-{nuovo_progressivo:06d}"
                    else:
                        numero_doc_finale = testata_data.get('numero_documento_manuale', 'ERRORE_NUM_MANCANTE')

                    # === FINE LOGICA DI NUMERAZIONE AUTOMATICA ===



                    nuova_testata = DocumentoTestata.objects.create(
                        tipo_doc=tipo_doc,
                        anagrafica=anagrafica,
                        data_documento=date.fromisoformat(testata_data['data_documento']),
                        numero_documento=numero_doc_finale, # <-- USIAMO IL NUMERO GENERATO
                        modalita_pagamento=modalita_pagamento,
                        cantiere=cantiere,
                        note=testata_data['note'],
                        imponibile=sum(Decimal(r['imponibile_riga']) for r in righe_data),
                        iva=sum(Decimal(r['iva_riga']) for r in righe_data),
                        totale=totale_documento,
                        stato=DocumentoTestata.Stato.CONFERMATO,
                        created_by=request.user,
                        updated_by=request.user
                    )
                    
                    for riga in righe_data:
                        DocumentoRiga.objects.create(
                            testata=nuova_testata, descrizione=riga['descrizione'], quantita=riga['quantita'],
                            prezzo_unitario=riga['prezzo_unitario'], aliquota_iva_id=riga['aliquota_iva_id'],
                            imponibile_riga=riga['imponibile_riga'], iva_riga=riga['iva_riga']
                        )
                    
                    for scadenza in scadenze_data:
                        Scadenza.objects.create(
                            documento=nuova_testata, anagrafica=anagrafica,
                            data_scadenza=date.fromisoformat(scadenza['data_scadenza']),
                            importo_rata=scadenza['importo_rata'],
                            tipo_scadenza=Scadenza.Tipo.INCASSO if 'V' in nuova_testata.tipo_doc else Scadenza.Tipo.PAGAMENTO
                        )
                    
                    clear_doc_wizard_session(request.session)
                    messages.success(request, f"Documento {nuova_testata} creato con successo!")
                    return redirect(nuova_testata.get_absolute_url())

        elif form.is_valid():
            nuova_scadenza = form.cleaned_data
            scadenze_data.append({
                'importo_rata': str(nuova_scadenza['importo_rata']),
                'data_scadenza': nuova_scadenza['data_scadenza'].isoformat(),
            })

            request.session['doc_scadenze_data'] = scadenze_data
            return redirect(reverse('documento_create_step3_scadenze'))
    else: # GET
        form = ScadenzaWizardForm(initial=get_scadenza_initial_data(testata_data, scadenze_data, residuo_da_scadenzare))

    context = {
        'form': form, 'totale_documento': totale_documento,
        'scadenze_inserite': scadenze_data, 'totale_scadenze': totale_scadenze,
        'residuo_da_scadenzare': residuo_da_scadenzare
    }
    return render(request, 'gestionale/documento_create_step3.html', context)

# === NUOVA FUNZIONE HELPER ===
# Per non ripetere il codice, ho estratto la logica di calcolo dei dati iniziali.
def get_scadenza_initial_data(testata_data, scadenze_data, residuo_da_scadenzare):
    data_proposta = None
    if scadenze_data:
        ultima_data = date.fromisoformat(scadenze_data[-1]['data_scadenza'])
        nuovo_mese, nuovo_anno = (ultima_data.month + 1, ultima_data.year)
        if nuovo_mese > 12:
            nuovo_mese, nuovo_anno = (1, nuovo_anno + 1)
        try:
            data_proposta = ultima_data.replace(year=nuovo_anno, month=nuovo_mese)
        except ValueError:
            data_proposta = ultima_data.replace(year=nuovo_anno, month=nuovo_mese, day=28)
    else:
        giorni_scadenza = 30
        try:
            mp_id = testata_data.get('modalita_pagamento_id')
            if mp_id:
                mp = ModalitaPagamento.objects.get(pk=mp_id)
                giorni_scadenza = mp.giorni_scadenza
        except ModalitaPagamento.DoesNotExist:
            pass
        data_documento = date.fromisoformat(testata_data['data_documento'])
        data_proposta = data_documento + timedelta(days=giorni_scadenza)

    return {
        'importo_rata': residuo_da_scadenzare,
        'data_scadenza': data_proposta.strftime('%Y-%m-%d')
    }

class DocumentoDeleteView(TenantRequiredMixin, RoleRequiredMixin, View):
    """
    Gestisce l'eliminazione sicura di un DocumentoTestata con un meccanismo
    di doppia conferma.
    L'eliminazione è permessa solo agli admin e solo se non esistono
    pagamenti collegati al documento.
    """
    allowed_roles = ['admin']
    template_name = 'gestionale/documento_confirm_delete.html'

    def get(self, request, *args, **kwargs):
        documento = get_object_or_404(DocumentoTestata, pk=kwargs['pk'])
        
        # Regola di Sicurezza: Controlla se esistono pagamenti
        has_pagamenti = PrimaNota.objects.filter(scadenza_collegata__documento=documento).exists()
        
        context = {
            'documento': documento,
            'has_pagamenti': has_pagamenti,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        documento = get_object_or_404(DocumentoTestata, pk=kwargs['pk'])
        
        # Regola di Sicurezza: Ricontrolla anche nella richiesta POST per massima sicurezza
        if PrimaNota.objects.filter(scadenza_collegata__documento=documento).exists():
            messages.error(request, "Impossibile eliminare il documento: sono presenti pagamenti/incassi collegati.")
            return redirect('documento_detail', pk=documento.pk)
        
        # Validazione della doppia conferma
        conferma_checkbox = request.POST.get('conferma_checkbox')
        conferma_testo = request.POST.get('conferma_testo')

        if not conferma_checkbox:
            messages.error(request, "Devi spuntare la casella di conferma per procedere.")
            return redirect('documento_delete', pk=documento.pk)

        if conferma_testo != 'ELIMINA': # Usiamo una parola in maiuscolo per aumentare l'attenzione
            messages.error(request, "La parola di conferma digitata non è corretta. Riprova.")
            return redirect('documento_delete', pk=documento.pk)

        # Se tutti i controlli sono superati, procedi con l'eliminazione
        try:
            with transaction.atomic():
                # Grazie a on_delete=CASCADE, Django cancellerà automaticamente
                # righe, scadenze e movimenti di primanota collegati.
                documento_info = f"{documento.get_tipo_doc_display()} N. {documento.numero_documento}"
                documento.delete()
                messages.success(request, f"Il documento '{documento_info}' e tutti i suoi dati collegati sono stati eliminati con successo.")
                return redirect('documento_list')
        except Exception as e:
            messages.error(request, f"Si è verificato un errore imprevisto durante l'eliminazione: {e}")
            return redirect('documento_detail', pk=kwargs['pk'])


# ==============================================================================
# === VISTE API (per richieste AJAX)                                        ===
# ==============================================================================

@login_required
@tenant_required
def get_anagrafiche_by_tipo(request):
    tipo_documento = request.GET.get('tipo_doc')
    anagrafiche_qs = Anagrafica.objects.none()

    # Logica per i documenti di VENDITA
    if tipo_documento in [DocumentoTestata.TipoDoc.FATTURA_VENDITA, DocumentoTestata.TipoDoc.NOTA_CREDITO_VENDITA]:
        anagrafiche_qs = Anagrafica.objects.filter(tipo=Anagrafica.Tipo.CLIENTE, attivo=True)
    
    # === NUOVA LOGICA PER I DOCUMENTI DI ACQUISTO ===
    elif tipo_documento in [DocumentoTestata.TipoDoc.FATTURA_ACQUISTO, DocumentoTestata.TipoDoc.NOTA_CREDITO_ACQUISTO]:
        anagrafiche_qs = Anagrafica.objects.filter(tipo=Anagrafica.Tipo.FORNITORE, attivo=True)

    data = [{'id': anag.pk, 'text': str(anag)} for anag in anagrafiche_qs]
    return JsonResponse({'results': data})

# ==============================================================================
# === VISTE DOCUMENTI (Dettaglio/Partitario)                              ===
# ==============================================================================


def get_documento_dettaglio_context(pk):
    """
    Funzione helper che recupera tutti i dati necessari per la vista
    di dettaglio di un documento, ma SENZA applicare la paginazione.
    """
    documento = get_object_or_404(DocumentoTestata, pk=pk)

    # Recupera i queryset completi e annota le scadenze con l'importo già pagato
    # Codice OTTIMIZZATO
    scadenze_qs = Scadenza.objects.filter(documento=documento).annotate(
        pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField()),
        # Aggiungiamo il calcolo del residuo direttamente nella query
        residuo=models.F('importo_rata') - models.F('pagato')
    ).order_by('data_scadenza')

    cronologia_pagamenti_qs = PrimaNota.objects.filter(
        scadenza_collegata__documento=documento
    ).select_related('scadenza_collegata', 'conto_finanziario').order_by('-data_registrazione')


    # Calcola il saldo totale del documento
    saldo_residuo = documento.totale - sum(s.pagato for s in scadenze_qs)

    # Restituisce i dati pronti per essere paginati dalla vista
    return {
        'documento': documento,
        'scadenze_qs': scadenze_qs,
        'cronologia_pagamenti_qs': cronologia_pagamenti_qs,
        'saldo_residuo': saldo_residuo
    }


class DocumentoDetailView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Mostra la vista di dettaglio completa di un singolo documento,
    gestendo due paginazioni indipendenti per scadenze e pagamenti.
    """
    template_name = 'gestionale/documento_detail.html'

    def get(self, request, *args, **kwargs):
        # Usa la funzione helper per ottenere i dati di base e i queryset completi
        context = get_documento_dettaglio_context(kwargs['pk'])

        # Gestione della paginazione per le SCADENZE
        pagina_scadenze_num = request.GET.get('pagina_scadenze', 1)
        paginator_scadenze = Paginator(context['scadenze_qs'], 5) # 5 elementi per pagina
        pagina_scadenze_obj = paginator_scadenze.get_page(pagina_scadenze_num)

        # Gestione della paginazione per i PAGAMENTI
        pagina_pagamenti_num = request.GET.get('pagina_pagamenti', 1)
        paginator_pagamenti = Paginator(context['cronologia_pagamenti_qs'], 10) # 10 elementi per pagina
        pagina_pagamenti_obj = paginator_pagamenti.get_page(pagina_pagamenti_num)

        # Aggiorna il contesto da passare al template
        context['scadenze'] = pagina_scadenze_obj
        context['cronologia_pagamenti'] = pagina_pagamenti_obj
        context['pagamento_form'] = PagamentoForm()

        # Rimuove i queryset non più necessari per pulizia
        context.pop('scadenze_qs')
        context.pop('cronologia_pagamenti_qs')
        context['conti_finanziari'] = ContoFinanziario.objects.filter(attivo=True)

        # Renderizza il template con il contesto finale
        return render(request, self.template_name, context)

# ==============================================================================
# === VISTE DOCUMENTI (LISTA + EXPORTS)                                     ===
# ==============================================================================

class DocumentoListView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Mostra l'elenco paginato e filtrabile di tutti i documenti.
    Contiene la logica di recupero dati riutilizzabile dagli export.
    """
    template_name = 'gestionale/documento_list.html'
    paginate_by = 15

    def _get_filtered_data(self, request):
        """
        Metodo helper che centralizza il recupero e il filtraggio dei documenti.
        """
        filter_form = DocumentoFilterForm(request.GET or None)
        
        documenti_qs = DocumentoTestata.objects.select_related('anagrafica').order_by('data_documento', 'numero_documento')

        if filter_form.is_valid():
            tipo_doc = filter_form.cleaned_data.get('tipo_doc')
            if tipo_doc:
                documenti_qs = documenti_qs.filter(tipo_doc=tipo_doc)
            data_da = filter_form.cleaned_data.get('data_da')
            if data_da:
                documenti_qs = documenti_qs.filter(data_documento__gte=data_da)
            data_a = filter_form.cleaned_data.get('data_a')
            if data_a:
                documenti_qs = documenti_qs.filter(data_documento__lte=data_a)
        
        return documenti_qs, filter_form

    def get(self, request, *args, **kwargs):
        documenti_qs, filter_form = self._get_filtered_data(request)
        
        paginator = Paginator(documenti_qs, self.paginate_by)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'documenti': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
            'filter_form': filter_form,
        }
        return render(request, self.template_name, context)

class DocumentoListExportExcelView(DocumentoListView):
    """
    Esporta la lista filtrata dei documenti in formato Excel.
    """
    def get(self, request, *args, **kwargs):
        # Chiama il nostro metodo helper per ottenere i dati filtrati.
        documenti_qs, filter_form = self._get_filtered_data(request)
        
        tenant_name = request.session.get('active_tenant_name', 'GestionaleDjango')
        report_title = 'Elenco Documenti'
        filename_prefix = 'Elenco_Documenti'
        
        # Costruisce la stringa dei filtri applicati.
        filtri_attivi = []
        if filter_form.is_valid() and filter_form.cleaned_data:
            for name, value in filter_form.cleaned_data.items():
                if value:
                    label = filter_form.fields[name].label or name.title()
                    display_value = value
                    if hasattr(filter_form.fields[name], 'choices'):
                        display_value = dict(filter_form.fields[name].choices).get(value, value)
                    if isinstance(value, date):
                        display_value = value.strftime('%d/%m/%Y')
                    filtri_attivi.append(f"{label}: {display_value}")
        filtri_str = " | ".join(filtri_attivi) if filtri_attivi else "Nessun filtro"
        
        headers = ["Tipo", "Numero", "Data", "Cliente/Fornitore", "Totale", "Stato"]
        data_rows = []
        for doc in documenti_qs:
            data_rows.append([
                doc.get_tipo_doc_display(),
                doc.numero_documento,
                doc.data_documento,
                doc.anagrafica.nome_cognome_ragione_sociale,
                doc.totale,
                doc.get_stato_display()
            ])
            
        report_sections = [{'title': 'Elenco Documenti', 'headers': headers, 'rows': data_rows}]
            
        return generate_excel_report(
            tenant_name, report_title, filtri_str, None, report_sections,
            filename_prefix=filename_prefix
        )

class DocumentoListExportPdfView(DocumentoListView):
    """
    Esporta la lista filtrata dei documenti in formato PDF.
    """
    def get(self, request, *args, **kwargs):
        # Chiama il nostro metodo helper per ottenere i dati filtrati.
        documenti_qs, filter_form = self._get_filtered_data(request)
        
        # Costruisce la stringa dei filtri applicati.
        filtri_attivi = []
        if filter_form.is_valid() and filter_form.cleaned_data:
            for name, value in filter_form.cleaned_data.items():
                if value:
                    label = filter_form.fields[name].label or name.title()
                    display_value = value
                    if hasattr(filter_form.fields[name], 'choices'):
                        display_value = dict(filter_form.fields[name].choices).get(value, value)
                    if isinstance(value, date):
                        display_value = value.strftime('%d/%m/%Y')
                    filtri_attivi.append(f"{label}: {display_value}")
        filtri_str = " | ".join(filtri_attivi) if filtri_attivi else "Nessun filtro"

        context = {
            'tenant_name': request.session.get('active_tenant_name', 'GestionaleDjango'),
            'report_title': 'Elenco Documenti',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            'documenti': documenti_qs,
        }
        
        return generate_pdf_report(
            request,
            'gestionale/documento_list_pdf.html', 
            context
        )

class DocumentoDetailExportPdfView(TenantRequiredMixin, RoleRequiredMixin, View):
   allowed_roles = ['admin', 'contabile', 'visualizzatore']
   """
   Gestisce la generazione e il download del PDF per il dettaglio di un documento.
   """
   def get(self, request, *args, **kwargs):
       # 1. Recupera tutti i dati necessari usando la funzione helper esistente.
       # Nota: la funzione helper restituisce i queryset completi, non paginati.
       context = get_documento_dettaglio_context(kwargs['pk'])

       # 2. Prepara il contesto specifico per il template PDF.
       context.update({
           'tenant_name': request.session.get('active_tenant_name', 'Gestionale'),
           'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
           # Passiamo i queryset completi al template PDF
           'scadenze': context['scadenze_qs'],
           'cronologia_pagamenti': context['cronologia_pagamenti_qs'],
       })

       # 3. Costruisce il nome del file secondo le specifiche.
       doc = context['documento']
       timestamp_str = timezone.now().strftime('%Y%m%d%H%M%S')
       doc_year = doc.data_documento.year
       # Pulisce il numero documento da eventuali caratteri non validi per un nome file
       safe_doc_number = doc.numero_documento.replace('/', '_').replace('\\', '_')

       filename = f"{timestamp_str}-Dettaglio_Ft_{doc_year}_{safe_doc_number}.pdf"

       # 4. Chiama la funzione di utility per generare il PDF.
       return generate_pdf_report(
           request,
           'gestionale/documento_detail_pdf.html',
           context,
           filename=filename  # Passiamo il nome file personalizzato
       )

# ==============================================================================
# === VISTE PARTITARIO ANAGRAFICA (DETAIL + EXPORTS)                        ===
# ==============================================================================

class AnagraficaDetailView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Gestisce la visualizzazione del partitario e contiene la logica
    di recupero dati riutilizzata dagli export.
    """
    template_name = 'gestionale/anagrafica_detail.html'
    
    def _get_partitario_data(self, request, anagrafica_pk):
        """
        Metodo helper che recupera e filtra TUTTI i dati per il partitario,
        inclusa la logica per il calcolo del saldo precedente.
        """
        anagrafica = get_object_or_404(Anagrafica, pk=anagrafica_pk)
        filter_form = PartitarioFilterForm(request.GET or None)
        
        # Inizializza le variabili
        saldo_precedente = Decimal('0.00')
        data_da = None

        # Queryset di base (non filtrati per data ancora)
        tutti_i_documenti = DocumentoTestata.objects.filter(anagrafica=anagrafica, stato=DocumentoTestata.Stato.CONFERMATO)
        tutti_i_movimenti = PrimaNota.objects.filter(anagrafica=anagrafica)

        # Applica i filtri e calcola il saldo precedente
        if filter_form.is_valid():
            data_da = filter_form.cleaned_data.get('data_da')
            if data_da:
                # 1. CALCOLO SALDO PRECEDENTE
                documenti_precedenti = tutti_i_documenti.filter(data_documento__lt=data_da)
                movimenti_precedenti = tutti_i_movimenti.filter(data_registrazione__lt=data_da)
                
                esposizione_precedente = sum(d.totale if 'V' in d.tipo_doc else -d.totale for d in documenti_precedenti)
                netto_movimenti_precedente = sum(m.importo if m.tipo_movimento == 'E' else -m.importo for m in movimenti_precedenti)
                saldo_precedente = esposizione_precedente - netto_movimenti_precedente

            data_a = filter_form.cleaned_data.get('data_a')
        
        # 2. FILTRA I DATI PER IL PERIODO SELEZIONATO
        documenti_periodo = tutti_i_documenti
        scadenze_periodo = Scadenza.objects.filter(anagrafica=anagrafica, stato__in=[Scadenza.Stato.APERTA, Scadenza.Stato.PARZIALE])
        movimenti_periodo = tutti_i_movimenti

        if data_da:
            documenti_periodo = documenti_periodo.filter(data_documento__gte=data_da)
            scadenze_periodo = scadenze_periodo.filter(data_scadenza__gte=data_da)
            movimenti_periodo = movimenti_periodo.filter(data_registrazione__gte=data_da)
        if 'data_a' in locals() and data_a:
            documenti_periodo = documenti_periodo.filter(data_documento__lte=data_a)
            scadenze_periodo = scadenze_periodo.filter(data_scadenza__lte=data_a)
            movimenti_periodo = movimenti_periodo.filter(data_registrazione__lte=data_a)
        
        # 3. APPLICA ORDINAMENTI E ANNOTAZIONI AI DATI DEL PERIODO
        documenti = documenti_periodo.order_by('-data_documento')
        scadenze_aperte = scadenze_periodo.annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField()),
            residuo=models.F('importo_rata') - models.F('pagato')
        ).order_by('data_scadenza')
        movimenti = movimenti_periodo.order_by('-data_registrazione')
        
        # 4. CALCOLA I KPI DEL PERIODO
        esposizione_periodo = sum(doc.totale if 'V' in doc.tipo_doc else -doc.totale for doc in documenti)
        netto_movimenti_periodo = sum(mov.importo if mov.tipo_movimento == 'E' else -mov.importo for mov in movimenti)
        
        # Il saldo finale è il saldo precedente + i movimenti del periodo
        saldo_finale = saldo_precedente + esposizione_periodo - netto_movimenti_periodo
        
        return {
            "anagrafica": anagrafica, "filter_form": filter_form,
            "documenti": documenti, "scadenze_aperte": scadenze_aperte,
            "movimenti": movimenti,
            # Passiamo sia i valori del periodo che quelli totali/precedenti
            "esposizione_documenti": esposizione_periodo,
            "netto_movimenti": netto_movimenti_periodo,
            "saldo_finale": saldo_finale,
            "saldo_precedente": saldo_precedente,
            "data_da_filtrata": data_da # Ci serve per la visualizzazione condizionale
        }

    def get(self, request, *args, **kwargs):
        """Metodo per la visualizzazione della pagina HTML."""
        partitario_data = self._get_partitario_data(request, kwargs['pk'])
        context = {}
        context.update(partitario_data)
        paginator_scadenze = Paginator(partitario_data['scadenze_aperte'], 5)
        page_number_scadenze = request.GET.get('pagina_scadenze', 1)
        context['scadenze_aperte'] = paginator_scadenze.get_page(page_number_scadenze)
        paginator_movimenti = Paginator(partitario_data['movimenti'], 5)
        page_number_movimenti = request.GET.get('pagina_movimenti', 1)
        context['movimenti'] = paginator_movimenti.get_page(page_number_movimenti)
        context['pagamento_form'] = PagamentoForm()
        context['conti_finanziari'] = ContoFinanziario.objects.filter(attivo=True)
        return render(request, self.template_name, context)

class AnagraficaPartitarioExportExcelView(AnagraficaDetailView):
    """
    Gestisce la creazione e il download di un report Excel per il partitario
    completo e filtrato di una specifica anagrafica.

    Eredita da AnagraficaDetailView per riutilizzare il metodo helper
    _get_partitario_data, garantendo che i dati esportati siano
    esattamente quelli visualizzati nella pagina HTML.
    """
    def get(self, request, *args, **kwargs):
        # 1. RECUPERO DEI DATI
        # Chiamiamo il nostro metodo helper centralizzato. Questo ci restituisce
        # un dizionario con tutti i dati già filtrati in base ai parametri GET
        # presenti nella richiesta.
        partitario_data = self._get_partitario_data(request, kwargs['pk'])
        anagrafica = partitario_data['anagrafica']

        # 2. PREPARAZIONE DEI DATI PER IL REPORT
        tenant_name = request.session.get('active_tenant_name', 'GestionaleDjango')
        
        # Titolo visualizzato DENTRO il report.
        report_title = f"Partitario {anagrafica.get_tipo_display()}: {anagrafica.nome_cognome_ragione_sociale}"
        
        # Prepariamo un nome di file pulito, rimuovendo caratteri non validi.
        safe_anag_name = "".join(c for c in anagrafica.nome_cognome_ragione_sociale if c.isalnum() or c in " _-").rstrip()
        filename_prefix = f"Partitario_{safe_anag_name}"
        
        # Usiamo la nostra utility per costruire la stringa dei filtri applicati.
        filtri_str = build_filters_string(partitario_data['filter_form'])

        # Prepariamo il dizionario dei KPI.
        # La struttura dei KPI cambia se è stato applicato un filtro 'Data Da'.
        kpi_report = {}
        if partitario_data.get('data_da_filtrata'):
            # Se il filtro è attivo, mostriamo il saldo precedente e i movimenti del periodo.
            data_da_str = partitario_data['data_da_filtrata'].strftime('%d/%m/%Y')
            kpi_report[f"Saldo al {data_da_str} (prec.)"] = partitario_data['saldo_precedente']
            kpi_report['Esposizione Documenti (nel periodo)'] = partitario_data['esposizione_documenti']
            kpi_report['Netto Movimenti (nel periodo)'] = partitario_data['netto_movimenti']
        else:
            # Altrimenti, mostriamo i totali complessivi.
            kpi_report['Esposizione Documenti'] = partitario_data['esposizione_documenti']
            kpi_report['Netto Incassato/Pagato'] = partitario_data['netto_movimenti']
        
        # Il saldo finale è sempre presente.
        kpi_report['Saldo Aperto Finale'] = partitario_data['saldo_finale']

        # 3. COSTRUZIONE DELLE SEZIONI DEL REPORT
        # Creiamo una lista che conterrà i dizionari di ogni sezione.
        report_sections = []

        # Sezione 1: Storico Documenti
        doc_headers = ["Data", "Tipo", "Numero", "Totale"]
        doc_rows = []
        for doc in partitario_data['documenti']:
            doc_rows.append([
                doc.data_documento,
                doc.get_tipo_doc_display(),
                doc.numero_documento,
                doc.totale
            ])
        report_sections.append({
            'title': 'Storico Documenti (Confermati)', 
            'headers': doc_headers, 
            'rows': doc_rows
        })

        # Sezione 2: Scadenziario Aperto
        scad_headers = ["Data Scad.", "Rif. Doc.", "Tipo", "Importo Rata", "Residuo", "Stato Rata"]
        scad_rows = []
        for scadenza in partitario_data['scadenze_aperte']:
            scad_rows.append([
                scadenza.data_scadenza,
                scadenza.documento.numero_documento,
                scadenza.get_tipo_scadenza_display(),
                scadenza.importo_rata,
                scadenza.residuo,
                scadenza.get_stato_display()
            ])
        report_sections.append({
            'title': 'Scadenziario Aperto', 
            'headers': scad_headers, 
            'rows': scad_rows
        })

        # Sezione 3: Cronologia Movimenti
        mov_headers = ["Data", "Descrizione", "Importo", "Conto"]
        mov_rows = []
        for movimento in partitario_data['movimenti']:
            mov_rows.append([
                movimento.data_registrazione,
                movimento.descrizione,
                movimento.importo * (1 if movimento.tipo_movimento == 'E' else -1),
                movimento.conto_finanziario.nome_conto
            ])
        report_sections.append({
            'title': 'Cronologia Movimenti', 
            'headers': mov_headers, 
            'rows': mov_rows
        })

        # 4. CHIAMATA ALLA FUNZIONE DI UTILITY
        # Passiamo tutti i dati preparati alla nostra funzione centralizzata.
        return generate_excel_report(
            tenant_name=tenant_name,
            report_title=report_title,
            filters_string=filtri_str,
            kpi_data=kpi_report,
            report_sections=report_sections,
            filename_prefix=filename_prefix
        )  

class AnagraficaPartitarioExportPdfView(AnagraficaDetailView):
    """
    Gestisce la creazione e il download di un report PDF per il partitario
    completo e filtrato di una specifica anagrafica.

    Eredita da AnagraficaDetailView per riutilizzare il metodo helper
    _get_partitario_data, garantendo che i dati esportati siano
    esattamente quelli visualizzati nella pagina HTML.
    """
    def get(self, request, *args, **kwargs):
        # 1. RECUPERO DEI DATI
        # Chiamiamo il nostro metodo helper centralizzato. Questo ci restituisce
        # un dizionario con tutti i dati già filtrati in base ai parametri GET.
        partitario_data = self._get_partitario_data(request, kwargs['pk'])
        anagrafica = partitario_data['anagrafica']

        # 2. PREPARAZIONE DEI DATI PER IL TEMPLATE PDF
        
        # Usiamo la nostra utility per costruire la stringa dei filtri applicati.
        filtri_str = build_filters_string(partitario_data['filter_form'])
        
        # Prepariamo il dizionario di contesto che verrà passato al template HTML
        # che poi WeasyPrint convertirà in PDF.
        context = {
            'tenant_name': request.session.get('active_tenant_name', 'GestionaleDjango'),
            'report_title': f"Partitario {anagrafica.get_tipo_display()}",
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            # Usiamo lo "splat operator" (**) per inserire tutte le coppie chiave-valore
            # del dizionario 'partitario_data' direttamente nel contesto.
            # In questo modo, il template avrà accesso a 'anagrafica', 'documenti',
            # 'scadenze_aperte', 'saldo_precedente', etc.
            **partitario_data
        }
        
        # 3. CHIAMATA ALLA FUNZIONE DI UTILITY
        # Passiamo la request, il nome del template da renderizzare e il contesto
        # alla nostra funzione centralizzata che gestirà la creazione del PDF.
        return generate_pdf_report(
            request,
            'gestionale/partitario_pdf_template.html', 
            context
        )


class AnagraficaListExportExcelView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']

    def get(self, request, *args, **kwargs):
        # Riutilizziamo la stessa logica di filtraggio
        filter_form = AnagraficaFilterForm(request.GET)
        query = Anagrafica.objects.all().order_by('nome_cognome_ragione_sociale')
        
        if filter_form.is_valid():
            q = filter_form.cleaned_data.get('q')
            tipo = filter_form.cleaned_data.get('tipo')
            attivo_str = filter_form.cleaned_data.get('attivo')
            if q:
                query = query.filter(Q(nome_cognome_ragione_sociale__icontains=q) | Q(citta__icontains=q) | Q(p_iva__icontains=q))
            if tipo:
                query = query.filter(tipo=tipo)
            if attivo_str:
                query = query.filter(attivo=(attivo_str == 'true'))

        report_sections = [{
            'title': 'Lista Anagrafiche',
            'headers': ['Codice', 'Nome/Ragione Sociale', 'Tipo', 'P.IVA', 'C.F.', 'Città', 'Stato'],
            'rows': [[a.codice, a.nome_cognome_ragione_sociale, a.get_tipo_display(), a.p_iva, a.codice_fiscale, a.citta, "Attivo" if a.attivo else "Non Attivo"] for a in query]
        }]
        
        return generate_excel_report(
            tenant_name=request.session.get('active_tenant_name', 'N/A'),
            report_title="Lista Anagrafiche",
            filters_string=build_filters_string(filter_form),
            kpi_data=None,
            report_sections=report_sections,
            filename_prefix="anagrafiche"
        )

class AnagraficaListExportPdfView(TenantRequiredMixin, RoleRequiredMixin, View):
    """
    Esporta la lista filtrata delle anagrafiche in formato PDF.
    """
    allowed_roles = ['admin', 'contabile', 'visualizzatore']

    def get(self, request, *args, **kwargs):
        # 1. Inizializziamo il form con i parametri GET, proprio come nella vista elenco
        filter_form = AnagraficaFilterForm(request.GET)
        
        # 2. Partiamo dal queryset di base
        query = Anagrafica.objects.all().order_by('nome_cognome_ragione_sociale')
        
        # 3. Applichiamo la stessa, identica logica di filtraggio
        if filter_form.is_valid():
            q = filter_form.cleaned_data.get('q')
            tipo = filter_form.cleaned_data.get('tipo')
            attivo_str = filter_form.cleaned_data.get('attivo')

            if q:
                query = query.filter(
                    Q(nome_cognome_ragione_sociale__icontains=q) |
                    Q(citta__icontains=q) |
                    Q(p_iva__icontains=q)
                )
            
            if tipo:
                query = query.filter(tipo=tipo)
            
            if attivo_str:
                attivo_bool = (attivo_str == 'true')
                query = query.filter(attivo=attivo_bool)

        # 4. Prepariamo il contesto per il template PDF
        context = {
            'tenant_name': request.session.get('active_tenant_name'),
            'report_title': 'Elenco Anagrafiche',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'anagrafiche': query,  # Usiamo il queryset filtrato
            'filters_string': build_filters_string(filter_form), # Aggiungiamo la stringa dei filtri
            'request': request, # WeasyPrint ha bisogno del request per gli URL assoluti delle risorse statiche
        }
        
        # 5. Chiamiamo la funzione di generazione del report
        return generate_pdf_report(
            request,
            'gestionale/anagrafica_list_pdf.html', 
            context
        )


# ==============================================================================
# === VISTE PAGAMENTI                                                       ===
# ==============================================================================

class RegistraPagamentoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile']
    """
    Gestisce la registrazione di un pagamento/incasso per una scadenza.
    """
    def post(self, request, *args, **kwargs):
        form = PagamentoForm(request.POST)
        redirect_url = request.META.get('HTTP_REFERER', reverse('dashboard'))

        if form.is_valid():
            scadenza = get_object_or_404(Scadenza, pk=form.cleaned_data['scadenza_id'])
            importo_pagato = form.cleaned_data['importo_pagato']
            
            # --- PRIMA CORREZIONE QUI ---
            # Aggiungiamo output_field=models.DecimalField() a Coalesce
            query_pagato = Scadenza.objects.filter(pk=scadenza.pk).annotate(
                pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField())
            ).first()
            pagato_precedente = query_pagato.pagato if query_pagato else Decimal('0.00')
            residuo_attuale = scadenza.importo_rata - pagato_precedente

            if importo_pagato > residuo_attuale:
                messages.error(request, f"L'importo inserito (€{importo_pagato}) supera il residuo (€{residuo_attuale:.2f}).")
                return redirect(redirect_url)

            with transaction.atomic():
                if scadenza.tipo_scadenza == Scadenza.Tipo.INCASSO:
                    causale, _ = Causale.objects.get_or_create(descrizione="INC. FT. CLI.")
                    tipo_movimento = PrimaNota.TipoMovimento.ENTRATA
                else:
                    causale, _ = Causale.objects.get_or_create(descrizione="PAG. FT. FORN.")
                    tipo_movimento = PrimaNota.TipoMovimento.USCITA
                
                PrimaNota.objects.create(
                    data_registrazione=form.cleaned_data['data_pagamento'],
                    descrizione=f"{causale.descrizione} - Doc. {scadenza.documento.numero_documento} - {scadenza.anagrafica.nome_cognome_ragione_sociale}",
                    importo=importo_pagato,
                    tipo_movimento=tipo_movimento,
                    conto_finanziario=form.cleaned_data['conto_finanziario'],
                    causale=causale,
                    anagrafica=scadenza.anagrafica,
                    scadenza_collegata=scadenza,
                    created_by=self.request.user
                )

                # --- SECONDA CORREZIONE QUI ---
                # Riutilizziamo la stessa logica robusta.
                # Il nuovo totale pagato è la somma di quanto già c'era più l'ultimo pagamento.
                nuovo_totale_pagato = pagato_precedente + importo_pagato
                
                if nuovo_totale_pagato >= scadenza.importo_rata:
                    scadenza.stato = Scadenza.Stato.SALDATA
                else:
                    scadenza.stato = Scadenza.Stato.PARZIALE
                scadenza.save()
            
            messages.success(request, "Pagamento registrato con successo.")
        else:
            # Rendiamo il messaggio di errore più specifico
            error_string = ". ".join([f"{key}: {value[0]}" for key, value in form.errors.items()])
            messages.error(request, f"Errore nella compilazione del form. {error_string}")
            
        return redirect(redirect_url)
    
class PagamentoDeleteView(TenantRequiredMixin, DeleteView):
    """
    Gestisce l'eliminazione di un pagamento (record di PrimaNota).
    Mostra una pagina di conferma e, dopo l'eliminazione, ricalcola
    e aggiorna lo stato della scadenza collegata.
    """
    model = PrimaNota
    template_name = 'gestionale/pagamento_confirm_delete.html'
    
    def get_success_url(self):
        """
        Restituisce l'URL a cui reindirizzare dopo l'eliminazione.
        Torniamo alla pagina di dettaglio del documento a cui apparteneva il pagamento.
        """
        scadenza = self.object.scadenza_collegata
        if scadenza:
            return reverse_lazy('documento_detail', kwargs={'pk': scadenza.documento.pk})
        # Fallback nel caso (improbabile) in cui non ci sia una scadenza
        return reverse_lazy('primanota_list')

    def form_valid(self, form):
        """
        Questo metodo viene eseguito dopo che l'utente ha confermato l'eliminazione.
        """
        # self.object è il record di PrimaNota che stiamo per cancellare.
        scadenza_da_aggiornare = self.object.scadenza_collegata
        
        with transaction.atomic():
            # Eseguiamo l'eliminazione effettiva del pagamento.
            # Chiamiamo il metodo della classe base per farlo.
            response = super().form_valid(form)
            
            # Se la scadenza esiste ancora, ricalcoliamo il suo stato.
            if scadenza_da_aggiornare:
                # Ricarichiamo l'oggetto dal DB per essere sicuri dei dati.
                scadenza_da_aggiornare.refresh_from_db()
                
                # Calcoliamo il nuovo totale pagato.
                aggregato = scadenza_da_aggiornare.pagamenti.aggregate(
                    total=Coalesce(Sum('importo'), Value(0), output_field=models.DecimalField())
                )
                totale_pagato = aggregato['total']
                
                # Aggiorniamo lo stato.
                if totale_pagato <= 0:
                    scadenza_da_aggiornare.stato = Scadenza.Stato.APERTA
                elif totale_pagato < scadenza_da_aggiornare.importo_rata:
                    scadenza_da_aggiornare.stato = Scadenza.Stato.PARZIALE
                else:
                    scadenza_da_aggiornare.stato = Scadenza.Stato.SALDATA
                
                scadenza_da_aggiornare.save()

        messages.success(self.request, f"Pagamento N. {self.object.pk} eliminato con successo. Stato scadenza aggiornato.")
        return response

class PagamentoUpdateView(TenantRequiredMixin, UpdateView):
    """
    Gestisce la modifica di un pagamento esistente (record di PrimaNota).
    """
    model = PrimaNota
    form_class = PagamentoUpdateForm
    template_name = 'gestionale/pagamento_form.html'
    
    def get_success_url(self):
        """
        Torna alla pagina di dettaglio del documento di origine.
        """
        return reverse_lazy('documento_detail', kwargs={'pk': self.object.scadenza_collegata.documento.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Modifica Pagamento N. {self.object.pk}"
        return context

    def form_valid(self, form):
        """
        Salva le modifiche e ricalcola lo stato della scadenza.
        """
        with transaction.atomic():
            # Salva le modifiche al pagamento
            pagamento = form.save(commit=False)
            pagamento.updated_by = self.request.user # Anche se non abbiamo questo campo su PrimaNota
            pagamento.save()
            
            # Ricalcola e aggiorna lo stato della scadenza collegata
            scadenza = pagamento.scadenza_collegata
            scadenza.refresh_from_db()
            
            totale_pagato = scadenza.pagamenti.aggregate(
                total=Coalesce(Sum('importo'), Value(0), output_field=models.DecimalField())
            )['total']
            
            if totale_pagato <= 0:
                scadenza.stato = Scadenza.Stato.APERTA
            elif totale_pagato < scadenza.importo_rata:
                scadenza.stato = Scadenza.Stato.PARZIALE
            else:
                scadenza.stato = Scadenza.Stato.SALDATA
            scadenza.save()

        messages.success(self.request, f"Pagamento N. {self.object.pk} aggiornato con successo.")
        return super().form_valid(form)
 
# ==============================================================================
# === VISTE SCADENZIARIO                                                    ===
# ==============================================================================

class ScadenzarioListView(TenantRequiredMixin, View):
    """
    Gestisce la visualizzazione della dashboard dello Scadenziario.
    Questa classe contiene la logica principale per recuperare e filtrare i dati,
    che viene poi riutilizzata anche dalle viste di export.
    """
    template_name = 'gestionale/scadenzario_list.html'
    paginate_by = 15

    def _get_filtered_data(self, request):
        """
        Metodo helper ("privato") che centralizza tutta la logica di recupero dati.
        Viene chiamato sia dalla vista HTML che dalle viste di export.
        
        Restituisce:
        - scadenze_qs: Il queryset delle scadenze, già filtrato.
        - kpi: Un dizionario con i totali aggregati per i KPI.
        - filter_form: L'istanza del form dei filtri, popolata con i dati GET.
        - today: La data odierna.
        """
        from datetime import date
        today = date.today()

        # Inizializza il form con i parametri GET della richiesta (es. ?tipo=Incasso)
        filter_form = ScadenzarioFilterForm(request.GET or None)
        
        # Queryset di base: tutte le scadenze che non sono né 'Saldata' né 'Annullata'.
        # Usiamo select_related per ottimizzare le query, pre-caricando i dati
        # delle tabelle collegate Anagrafica e DocumentoTestata con un unico JOIN.
        scadenze_qs = Scadenza.objects.filter(
            stato__in=[Scadenza.Stato.APERTA, Scadenza.Stato.PARZIALE]
        ).select_related('anagrafica', 'documento').order_by('data_scadenza')

        # Applica i filtri al queryset se il form è stato inviato e i dati sono validi.
        if filter_form.is_valid():
            cleaned_data = filter_form.cleaned_data
            
            if cleaned_data.get('anagrafica'):
                scadenze_qs = scadenze_qs.filter(anagrafica=cleaned_data['anagrafica'])
            
            if cleaned_data.get('data_da'):
                scadenze_qs = scadenze_qs.filter(data_scadenza__gte=cleaned_data['data_da'])
            
            if cleaned_data.get('data_a'):
                scadenze_qs = scadenze_qs.filter(data_scadenza__lte=cleaned_data['data_a'])
            
            if cleaned_data.get('tipo'):
                scadenze_qs = scadenze_qs.filter(tipo_scadenza=cleaned_data['tipo'])
            
            if cleaned_data.get('stato') == 'scadute':
                scadenze_qs = scadenze_qs.filter(data_scadenza__lt=today)
            elif cleaned_data.get('stato') == 'a_scadere':
                scadenze_qs = scadenze_qs.filter(data_scadenza__gte=today)
        
        # Calcola i KPI aggregati sull'INTERO queryset GIA' FILTRATO.
        # Questa è una singola, efficiente query al database.
        kpi = scadenze_qs.aggregate(
            incassi_totali_rate=Coalesce(Sum('importo_rata', filter=models.Q(tipo_scadenza=Scadenza.Tipo.INCASSO)), Value(0), output_field=models.DecimalField()),
            pagamenti_totali_rate=Coalesce(Sum('importo_rata', filter=models.Q(tipo_scadenza=Scadenza.Tipo.PAGAMENTO)), Value(0), output_field=models.DecimalField()),
            incassi_pagati=Coalesce(Sum('pagamenti__importo', filter=models.Q(tipo_scadenza=Scadenza.Tipo.INCASSO)), Value(0), output_field=models.DecimalField()),
            pagamenti_pagati=Coalesce(Sum('pagamenti__importo', filter=models.Q(tipo_scadenza=Scadenza.Tipo.PAGAMENTO)), Value(0), output_field=models.DecimalField()),
            incassi_scaduti_rate=Coalesce(Sum('importo_rata', filter=models.Q(tipo_scadenza=Scadenza.Tipo.INCASSO, data_scadenza__lt=today)), Value(0), output_field=models.DecimalField()),
            pagamenti_scaduti_rate=Coalesce(Sum('importo_rata', filter=models.Q(tipo_scadenza=Scadenza.Tipo.PAGAMENTO, data_scadenza__lt=today)), Value(0), output_field=models.DecimalField()),
            incassi_scaduti_pagati=Coalesce(Sum('pagamenti__importo', filter=models.Q(tipo_scadenza=Scadenza.Tipo.INCASSO, data_scadenza__lt=today)), Value(0), output_field=models.DecimalField()),
            pagamenti_scaduti_pagati=Coalesce(Sum('pagamenti__importo', filter=models.Q(tipo_scadenza=Scadenza.Tipo.PAGAMENTO, data_scadenza__lt=today)), Value(0), output_field=models.DecimalField())
        )
        
        # Restituisce i dati pronti per essere usati.
        return scadenze_qs, kpi, filter_form, today

    def get(self, request, *args, **kwargs):
        """
        Metodo principale che gestisce la richiesta GET per la pagina HTML.
        Il suo compito è orchestrare la chiamata all'helper, gestire la paginazione
        e preparare il contesto finale per il template.
        """
        # 1. Chiama il metodo helper per ottenere tutti i dati di base.
        scadenze_qs, kpi_data, filter_form, today = self._get_filtered_data(request)

        # 2. Calcola i valori finali dei KPI dai dati aggregati.
        incassi_aperti = kpi_data['incassi_totali_rate'] - kpi_data['incassi_pagati']
        pagamenti_aperti = kpi_data['pagamenti_totali_rate'] - kpi_data['pagamenti_pagati']
        incassi_scaduti = kpi_data['incassi_scaduti_rate'] - kpi_data['incassi_scaduti_pagati']
        pagamenti_scaduti = kpi_data['pagamenti_scaduti_rate'] - kpi_data['pagamenti_scaduti_pagati']
        
        # 3. Applica la paginazione al queryset filtrato.
        paginator = Paginator(scadenze_qs, self.paginate_by)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # 4. Calcola il residuo per le sole scadenze della pagina corrente (per efficienza).
        ids_pagina_corrente = [s.id for s in page_obj]
        scadenze_pagina_con_pagato = Scadenza.objects.filter(pk__in=ids_pagina_corrente).annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField())
        )
        pagato_map = {s.id: s.pagato for s in scadenze_pagina_con_pagato}

        for s in page_obj:
            s.residuo = s.importo_rata - pagato_map.get(s.id, Decimal(0))

        # 5. Prepara il contesto da passare al template.
        context = {
            'page_obj': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'pagamento_form': PagamentoForm(), # Necessario per la modale
            'today': today,
            'filter_form': filter_form,
            'conti_finanziari': ContoFinanziario.objects.filter(attivo=True),
            'kpi': {
                'incassi_aperti': incassi_aperti,
                'pagamenti_aperti': pagamenti_aperti,
                'saldo_circolante': incassi_aperti - pagamenti_aperti,
                'incassi_scaduti': incassi_scaduti,
                'pagamenti_scaduti': pagamenti_scaduti,
            }
        }
        
        # 6. Renderizza il template con il contesto.
        return render(request, self.template_name, context)

class ScadenzarioExportExcelView(ScadenzarioListView):
    def get(self, request, *args, **kwargs):
        scadenze_qs, kpi_data, filter_form, today = self._get_filtered_data(request)
        
        tenant_name = request.session.get('active_tenant_name', 'GestionaleDjango')
        report_title = 'Report Scadenziario Aperto'
        filename_prefix = 'Scadenziario_Aperto'
        filtri_str = build_filters_string(filter_form)
        
        incassi_aperti = kpi_data['incassi_totali_rate'] - kpi_data['incassi_pagati']
        pagamenti_aperti = kpi_data['pagamenti_totali_rate'] - kpi_data['pagamenti_pagati']
        kpi_report = {
            'Incassi Aperti': incassi_aperti,
            'Pagamenti Aperti': pagamenti_aperti,
            'Saldo Circolante': incassi_aperti - pagamenti_aperti
        }
        
        # === CORREZIONE ===
        headers = ["Data Scad.", "Tipo", "Cliente/Fornitore", "Rif. Doc.", "Importo Rata", "Residuo", "Stato Rata"]
        data_rows = []
        scadenze_con_pagato = scadenze_qs.annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField())
        )
        for scadenza in scadenze_con_pagato:
            data_rows.append([
                scadenza.data_scadenza, scadenza.get_tipo_scadenza_display(),
                scadenza.anagrafica.nome_cognome_ragione_sociale, scadenza.documento.numero_documento,
                scadenza.importo_rata, scadenza.importo_rata - scadenza.pagato,
                scadenza.get_stato_display()
            ])
            
        report_sections = [{'title': 'Dettaglio Scadenze', 'headers': headers, 'rows': data_rows}]
        
        return generate_excel_report(
            tenant_name=tenant_name,
            report_title=report_title,
            filters_string=filtri_str,
            kpi_data=kpi_report,
            report_sections=report_sections, # Passiamo la lista di sezioni
            filename_prefix=filename_prefix
        )

class ScadenzarioExportPdfView(ScadenzarioListView):
    """
    Esporta i dati dello scadenziario in PDF, usando la utility centralizzata.
    """
    def get(self, request, *args, **kwargs):
        # 1. Recupera i dati filtrati e i KPI.
        scadenze_qs, kpi_data, filter_form, today = self._get_filtered_data(request)

        # 2. Prepara i dati per il contesto del template PDF.
        # Calcola il residuo per ogni scadenza.
        scadenze_con_pagato = scadenze_qs.annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=models.DecimalField())
        )
        for scadenza in scadenze_con_pagato:
            scadenza.residuo = scadenza.importo_rata - scadenza.pagato
            
        # Costruisci la stringa dei filtri.
        filtri_str = build_filters_string(filter_form)

        # Calcola i valori finali dei KPI.
        incassi_aperti = kpi_data['incassi_totali_rate'] - kpi_data['incassi_pagati']
        pagamenti_aperti = kpi_data['pagamenti_totali_rate'] - kpi_data['pagamenti_pagati']
        incassi_scaduti = kpi_data['incassi_scaduti_rate'] - kpi_data['incassi_scaduti_pagati']
        pagamenti_scaduti = kpi_data['pagamenti_scaduti_rate'] - kpi_data['pagamenti_scaduti_pagati']

        # 3. Crea il dizionario di contesto.
        context = {
            'tenant_name': request.session.get('active_tenant_name', 'GestionaleDjango'),
            'report_title': 'Report Scadenziario Aperto',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            'scadenze': scadenze_con_pagato,
            'kpi': {
                'incassi_aperti': incassi_aperti,
                'pagamenti_aperti': pagamenti_aperti,
                'saldo_circolante': incassi_aperti - pagamenti_aperti,
                'incassi_scaduti': incassi_scaduti,
                'pagamenti_scaduti': pagamenti_scaduti,
            }
        }
        
        # 4. Chiama la funzione di utility e restituisci il risultato.
        return generate_pdf_report(
            request, 
            'gestionale/scadenzario_pdf_template.html', 
            context
        )

# ==============================================================================
# === VISTE HR (Human Resources)                                            ===
# ==============================================================================

class DashboardHRView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Mostra la dashboard per la pianificazione e consuntivazione giornaliera
    delle risorse umane.
    """
    template_name = 'gestionale/dashboard_hr.html'

    def get(self, request, *args, **kwargs):
        # 1. Determina la data di riferimento (invariato)
        today = date.today()
        year = kwargs.get('year', today.year)
        month = kwargs.get('month', today.month)
        day = kwargs.get('day', today.day)
        try:
            data_riferimento = date(year, month, day)
        except ValueError:
            data_riferimento = today
            
        # 2. Recupera tutti i dipendenti attivi (invariato)
        dipendenti_attivi = Anagrafica.objects.filter(
            tipo=Anagrafica.Tipo.DIPENDENTE,
            attivo=True
        ).select_related('dettaglio_dipendente').order_by('nome_cognome_ragione_sociale')
        
        # 3. Recupera le attività del diario per la data di riferimento (invariato)
        diario_del_giorno = DiarioAttivita.objects.filter(data=data_riferimento).select_related('cantiere_pianificato', 'mezzo_pianificato')
        
        # 4. Mappiamo le attività ai dipendenti per un accesso veloce (invariato)
        mappa_diario = {d.dipendente_id: d for d in diario_del_giorno}
        
        # === INIZIO NUOVA LOGICA PER I CANTIERI ===
        # 1. Recupera il filtro per lo stato del cantiere dall'URL.
        #    Il default è 'APERTO', come da requisiti.
        stato_cantiere_filter = request.GET.get('stato_cantiere', Cantiere.Stato.APERTO)

        # 2. Queryset di base per i cantieri.
        cantieri_qs = Cantiere.objects.select_related('cliente').order_by('codice_cantiere')
        
        # 3. Applica il filtro, a meno che non sia 'TUTTI'.
        if stato_cantiere_filter and stato_cantiere_filter != 'TUTTI':
            cantieri_qs = cantieri_qs.filter(stato=stato_cantiere_filter)
        # === FINE NUOVA LOGICA PER I CANTIERI ===

        # 5. Combiniamo i dati, calcoliamo i KPI e le ore di default
        lista_dipendenti_con_stato = []
        kpi = {
            'presenti': 0, 'assenti': 0, 'assegnati': 0, 'liberi': 0,
            'totale_dipendenti': len(dipendenti_attivi)
        }
        
        for dip in dipendenti_attivi:
            attivita_giornaliera = mappa_diario.get(dip.pk)
            dip.attivita = attivita_giornaliera

            # === INIZIO LOGICA MANCANTE AGGIUNTA QUI ===
            # Calcoliamo le ore giornaliere di default per QUESTO dipendente
            dettaglio = dip.dettaglio_dipendente
            ore_default = "0.00" # Default di sicurezza
            if dettaglio and dettaglio.giorni_lavorativi_settimana and dettaglio.giorni_lavorativi_settimana > 0:
                ore_calcolate = dettaglio.ore_settimanali_contratto / dettaglio.giorni_lavorativi_settimana
                # Formattiamo come stringa con 2 decimali e punto
                ore_default = f"{ore_calcolate:.2f}".replace(',', '.')
            
            # Aggiungiamo il valore calcolato come un nuovo attributo all'oggetto dipendente
            dip.ore_default = ore_default
            # === FINE LOGICA MANCANTE ===

            lista_dipendenti_con_stato.append(dip)
            
            # Aggiorniamo i contatori dei KPI (invariato)
            if attivita_giornaliera:
                if attivita_giornaliera.stato_presenza == DiarioAttivita.StatoPresenza.PRESENTE:
                    kpi['presenti'] += 1
                elif attivita_giornaliera.stato_presenza in [DiarioAttivita.StatoPresenza.ASSENTE_G, DiarioAttivita.StatoPresenza.ASSENTE_I]:
                    kpi['assenti'] += 1
                elif attivita_giornaliera.cantiere_pianificato:
                    kpi['assegnati'] += 1
            else:
                kpi['liberi'] += 1
        
        kpi['disponibili'] = kpi['liberi']

        context = {
            'data_riferimento': data_riferimento,
            'giorno_precedente': data_riferimento - timedelta(days=1),
            'giorno_successivo': data_riferimento + timedelta(days=1),
            'dipendenti': lista_dipendenti_con_stato,
            'cantieri_disponibili': Cantiere.objects.filter(stato=Cantiere.Stato.APERTO),
            'mezzi_disponibili': MezzoAziendale.objects.filter(attivo=True),
            'attivita_form': DiarioAttivitaForm(),
            'kpi': kpi,
            'cantieri': cantieri_qs,
            'stati_cantiere': Cantiere.Stato.choices, # Passiamo le scelte per il dropdown
            'stato_cantiere_selezionato': stato_cantiere_filter,
        }
        
        return render(request, self.template_name, context)

class SalvaAttivitaDiarioView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile'] # Aggiunto AdminRequiredMixin per sicurezza
    """
    Gestisce il salvataggio (creazione o modifica) di una voce
    del DiarioAttivita tramite una richiesta POST dalla modale.
    Accessibile solo agli admin.
    """
    def post(self, request, *args, **kwargs):
        # 1. RECUPERO E VALIDAZIONE INPUT DI BASE
        data_str = request.POST.get('data')
        dipendente_id = request.POST.get('dipendente_id')
        
        # Costruiamo un URL di fallback a cui tornare in caso di errore grave
        fallback_url = reverse('dashboard_hr')

        # Se i dati essenziali mancano, non possiamo procedere.
        if not data_str or not dipendente_id:
            messages.error(request, "Dati mancanti o corrotti. Impossibile salvare.")
            return redirect(fallback_url)
            
        try:
            # Proviamo a convertire la stringa in una data.
            data = date.fromisoformat(data_str)
        except ValueError:
            # Se la stringa non è una data valida, blocchiamo l'operazione.
            messages.error(request, "Formato data non valido. Impossibile salvare.")
            return redirect(fallback_url)

        # Ora che 'data' è valida, possiamo costruire l'URL di reindirizzamento corretto.
        redirect_url = reverse('dashboard_hr_data', kwargs={'year': data.year, 'month': data.month, 'day': data.day})
        
        # 2. GET OR CREATE E VALIDAZIONE FORM (logica quasi invariata)
        attivita, created = DiarioAttivita.objects.get_or_create(
            data=data,
            dipendente_id=dipendente_id,
            defaults={'created_by': request.user}
        )
        
        form = DiarioAttivitaForm(request.POST, instance=attivita)

        if form.is_valid():
            # ... (tutta la logica di 'form_valid' rimane identica a prima) ...
            istanza_salvata = form.save(commit=False)
            istanza_salvata.updated_by = self.request.user
            
            if istanza_salvata.ore_ordinarie is None: istanza_salvata.ore_ordinarie = 0
            if istanza_salvata.ore_straordinarie is None: istanza_salvata.ore_straordinarie = 0
            
            if istanza_salvata.ore_ordinarie > 0 or istanza_salvata.ore_straordinarie > 0:
                if not istanza_salvata.stato_presenza:
                    istanza_salvata.stato_presenza = DiarioAttivita.StatoPresenza.PRESENTE
            
            stati_assenza = [DiarioAttivita.StatoPresenza.ASSENTE_G, DiarioAttivita.StatoPresenza.ASSENTE_I]
            if istanza_salvata.stato_presenza in stati_assenza:
                istanza_salvata.ore_ordinarie = 0
                istanza_salvata.ore_straordinarie = 0

            istanza_salvata.save()
            messages.success(request, f"Attività per {attivita.dipendente.nome_cognome_ragione_sociale} salvata con successo.")
        else:
            error_string = ". ".join([f"'{key}': {value[0]}" for key, value in form.errors.items()])
            messages.error(request, f"Errore nel salvataggio: {error_string}")
        
        return redirect(redirect_url)
    
# ==============================================================================
# === VISTE PRIMA NOTA                                                      ===
# ==============================================================================

class PrimaNotaListView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Gestisce la visualizzazione della dashboard di Prima Nota.
    Contiene la logica di recupero e filtraggio dati riutilizzabile dagli export.
    """
    template_name = 'gestionale/primanota_list.html'
    paginate_by = 15

    def _get_filtered_data(self, request):
        """
        Metodo helper che centralizza il recupero e il filtraggio dei movimenti.
        Restituisce il queryset filtrato e il form dei filtri.
        """
        filter_form = PrimaNotaFilterForm(request.GET or None)
        
        movimenti_qs = PrimaNota.objects.select_related(
            'conto_finanziario', 'causale', 'anagrafica', 'cantiere'
        ).order_by('-data_registrazione', '-pk')

        if filter_form.is_valid():
            cleaned_data = filter_form.cleaned_data
            if cleaned_data.get('descrizione'):
                movimenti_qs = movimenti_qs.filter(descrizione__icontains=cleaned_data['descrizione'])
            if cleaned_data.get('conto_finanziario'):
                movimenti_qs = movimenti_qs.filter(conto_finanziario=cleaned_data['conto_finanziario'])
            if cleaned_data.get('causale'):
                movimenti_qs = movimenti_qs.filter(causale=cleaned_data['causale'])
            if cleaned_data.get('data_da'):
                movimenti_qs = movimenti_qs.filter(data_registrazione__gte=cleaned_data['data_da'])
            if cleaned_data.get('data_a'):
                movimenti_qs = movimenti_qs.filter(data_registrazione__lte=cleaned_data['data_a'])
        
        # === INIZIO MODIFICA ===
        # Aggiungiamo il filtro per cantiere, che non fa parte del form
        # ma può essere passato come parametro GET da altre parti dell'applicazione.
        cantiere_id = request.GET.get('cantiere')
        if cantiere_id:
            movimenti_qs = movimenti_qs.filter(cantiere_id=cantiere_id)
        # === FINE MODIFICA ===

        return movimenti_qs, filter_form

    def get(self, request, *args, **kwargs):
        """
        Gestisce la richiesta GET per la pagina HTML.
        """
        movimenti_qs, filter_form = self._get_filtered_data(request)
        
        paginator = Paginator(movimenti_qs, self.paginate_by)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'movimenti': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
            'filter_form': filter_form,
        }
        return render(request, self.template_name, context)

class PrimaNotaDeleteView(TenantRequiredMixin, RoleRequiredMixin, DeleteView):
    allowed_roles = ['admin', 'contabile']
    """
    Gestisce l'eliminazione di un movimento di Prima Nota,
    mostrando un avviso speciale per i giroconti.
    """
    model = PrimaNota
    template_name = 'gestionale/primanota_confirm_delete.html'
    success_url = reverse_lazy('primanota_list')

    def form_valid(self, form):
        """
        Aggiunge un messaggio di successo dopo l'eliminazione.
        La logica di cancellazione del movimento collegato è nel modello.
        """
        messages.success(self.request, f"Movimento N. {self.object.pk} eliminato con successo.")
        return super().form_valid(form)

class PrimaNotaUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin', 'contabile']
    """
    Gestisce la modifica di un movimento di Prima Nota.
    Contiene la logica speciale per sincronizzare i due movimenti
    quando si modifica un giroconto.
    """
    model = PrimaNota
    form_class = PrimaNotaUpdateForm # Usa il form specifico per la modifica (per la data)
    template_name = 'gestionale/primanota_form.html'
    
    def dispatch(self, request, *args, **kwargs):
        """
        Controlla se il movimento è un giroconto prima di procedere.
        """
        # Recuperiamo l'oggetto prima di ogni altra cosa
        self.object = self.get_object()
        if self.object.movimento_collegato:
            messages.error(request, "I movimenti di giroconto non possono essere modificati. Si prega di eliminarli e ricrearli.")
            return redirect('primanota_list')

        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        """
        Restituisce l'URL a cui reindirizzare dopo un salvataggio riuscito.
        """
        return reverse_lazy('primanota_list')

    def get_context_data(self, **kwargs):
        """
        Prepara il contesto per il template, popolando i campi extra se necessario.
        """
        context = super().get_context_data(**kwargs)
        context['title'] = f"Modifica Movimento N. {self.object.pk}"
        
        # Passa l'ID della causale "Giroconto" per lo script JS.
        try:
            context['giroconto_causale_id'] = Causale.objects.get(descrizione__iexact="GIROCONTO").pk
        except Causale.DoesNotExist:
            context['giroconto_causale_id'] = None
            
        # Se stiamo modificando un giroconto, dobbiamo pre-popolare il campo
        # 'conto_destinazione' nel form, che non fa parte del modello.
        if self.object.movimento_collegato:
            form = context['form']
            # Il conto di "destinazione" è sempre il conto finanziario dell'altro movimento.
            form.initial['conto_destinazione'] = self.object.movimento_collegato.conto_finanziario

        return context
    
    def form_valid(self, form):
        """
        Questo metodo viene eseguito quando i dati inviati nel form sono validi.
        Contiene la logica di business per orchestrare il salvataggio.
        """
        # Recupera l'oggetto originale dal database, prima di qualsiasi modifica.
        movimento_originale = self.get_object()
        
        causale = form.cleaned_data.get('causale')
        is_giroconto = causale and causale.descrizione.upper() == 'GIROCONTO'

        # Eseguiamo sempre le modifiche in una transazione atomica per sicurezza.
        with transaction.atomic():
            # Ottieni l'oggetto aggiornato con i dati del form, ma non salvarlo ancora.
            movimento_aggiornato = form.save(commit=False)
            movimento_aggiornato.updated_by = self.request.user
            
            # --- CASO SPECIALE: GIROCONTO ---
            if is_giroconto:
                # Recupera il movimento collegato dal database.
                movimento_collegato = movimento_originale.movimento_collegato
                
                # Preserva il tipo_movimento del record principale, perché non arriva dal form.
                movimento_aggiornato.tipo_movimento = movimento_originale.tipo_movimento
                
                # Salva le modifiche al movimento principale.
                movimento_aggiornato.save()
                
                # Ora, se esiste un movimento collegato, sincronizzalo.
                if movimento_collegato:
                    # Recupera i conti dal form per aggiornare le descrizioni.
                    conto_origine = form.cleaned_data['conto_finanziario']
                    conto_destinazione = form.cleaned_data['conto_destinazione']

                    # Aggiorna i campi del movimento collegato con i nuovi valori.
                    movimento_collegato.data_registrazione = form.cleaned_data['data_registrazione']
                    movimento_collegato.importo = form.cleaned_data['importo']
                    movimento_collegato.updated_by = self.request.user

                    # Aggiorna la descrizione in modo speculare.
                    if movimento_collegato.tipo_movimento == PrimaNota.TipoMovimento.USCITA:
                        movimento_collegato.descrizione = f"GIROCONTO -> {conto_origine.nome_conto}"
                    else: # È un'entrata
                        movimento_collegato.descrizione = f"GIROCONTO <- {conto_destinazione.nome_conto}"
                    
                    movimento_collegato.save()
            
            # --- CASO NORMALE: MOVIMENTO STANDARD ---
            else:
                # Se non è un giroconto, salva semplicemente le modifiche.
                movimento_aggiornato.save()
        
        messages.success(self.request, f"Movimento N. {movimento_originale.pk} aggiornato con successo.")
        return HttpResponseRedirect(self.get_success_url())    

class PrimaNotaCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin', 'contabile']
    """
    Gestisce la creazione di un nuovo movimento di Prima Nota.
    Contiene la logica speciale per creare due movimenti atomici
    quando la causale selezionata è "GIROCONTO".
    """
    model = PrimaNota
    form_class = PrimaNotaForm # Usa il form che gestisce dinamicamente i campi
    template_name = 'gestionale/primanota_form.html'
    
    def get_success_url(self):
        """
        Restituisce l'URL a cui reindirizzare dopo un salvataggio riuscito.
        """
        return reverse_lazy('primanota_list')

    def get_context_data(self, **kwargs):
        """
        Aggiunge dati extra al contesto del template.
        """
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Movimento di Prima Nota'
        
        # Passiamo l'ID della causale "Giroconto" al template.
        # Questo ID verrà usato dallo script JavaScript per mostrare/nascondere
        # i campi del form in modo dinamico.
        try:
            context['giroconto_causale_id'] = Causale.objects.get(descrizione__iexact="GIROCONTO").pk
        except Causale.DoesNotExist:
            context['giroconto_causale_id'] = None
            
        return context

    def form_valid(self, form):
        """
        Questo metodo viene eseguito quando i dati inviati nel form sono validi.
        Contiene la logica di business per orchestrare il salvataggio.
        """
        causale = form.cleaned_data.get('causale')
        is_giroconto = causale and causale.descrizione.upper() == 'GIROCONTO'
        
        # --- CASO SPECIALE: GIROCONTO ---
        if is_giroconto:
            # Usiamo una transazione atomica per garantire l'integrità dei dati.
            # Se una delle operazioni di salvataggio fallisce, tutte le modifiche
            # al database vengono annullate.
            with transaction.atomic():
                conto_origine = form.cleaned_data['conto_finanziario']
                conto_destinazione = form.cleaned_data['conto_destinazione']
                importo = form.cleaned_data['importo']
                data_registrazione = form.cleaned_data['data_registrazione']
                
                # 1. Crea il movimento di USCITA
                # usiamo form.save(commit=False) per non salvare subito e poter
                # modificare i campi gestiti dal sistema.
                uscita = form.save(commit=False)
                uscita.created_by = self.request.user
                uscita.updated_by = self.request.user
                uscita.tipo_movimento = PrimaNota.TipoMovimento.USCITA
                uscita.descrizione = f"GIROCONTO -> {conto_destinazione.nome_conto}"
                uscita.save() # Primo salvataggio per ottenere un pk

                # 2. Crea il movimento di ENTRATA speculare
                entrata = PrimaNota.objects.create(
                    data_registrazione=data_registrazione,
                    descrizione=f"GIROCONTO <- {conto_origine.nome_conto}",
                    importo=importo,
                    tipo_movimento=PrimaNota.TipoMovimento.ENTRATA,
                    causale=causale,
                    conto_finanziario=conto_destinazione,
                    created_by=self.request.user,
                    movimento_collegato=uscita
                )
                
                # 3. Aggiorna il movimento di uscita per creare il legame bidirezionale
                uscita.movimento_collegato = entrata
                uscita.save()

            messages.success(self.request, "Giroconto registrato con successo.")

        # --- CASO NORMALE: MOVIMENTO STANDARD ---
        else:
            movimento = form.save(commit=False)
            movimento.created_by = self.request.user
            movimento.updated_by = self.request.user # Aggiungiamo anche updated_by per coerenza
            movimento.save()
            messages.success(self.request, "Movimento di prima nota creato con successo.")
            
        # Reindirizza alla pagina di successo in entrambi i casi.
        return HttpResponseRedirect(self.get_success_url())
    
class PrimaNotaListExportExcelView(PrimaNotaListView):
    """
    Esporta la lista filtrata dei movimenti di Prima Nota in formato Excel.
    Eredita da PrimaNotaListView per accedere al metodo _get_filtered_data.
    """
    def get(self, request, *args, **kwargs):
        # 1. RECUPERO DATI FILTRATI
        # Chiamiamo il nostro metodo helper per ottenere i dati esatti che l'utente vede.
        movimenti_qs, filter_form = self._get_filtered_data(request)
        
        # 2. PREPARAZIONE DEI DATI PER IL REPORT
        tenant_name = request.session.get('active_tenant_name', 'GestionaleDjango')
        report_title = 'Elenco Movimenti di Prima Nota'
        filename_prefix = 'Prima_Nota'
        
        # Costruisci la stringa dei filtri applicati
        filtri_attivi = []
        if filter_form.is_valid() and filter_form.cleaned_data:
            for name, value in filter_form.cleaned_data.items():
                if value:
                    label = filter_form.fields[name].label or name.title()
                    display_value = value
                    if hasattr(filter_form.fields[name], 'choices'):
                        display_value = dict(filter_form.fields[name].choices).get(value, value)
                    if isinstance(value, date):
                        display_value = value.strftime('%d/%m/%Y')
                    if isinstance(field, forms.ModelChoiceField): # 'field' non è definito, usiamo 'filter_form.fields[name]'
                        display_value = str(value)
                    filtri_attivi.append(f"{label}: {display_value}")
        filtri_str = " | ".join(filtri_attivi) if filtri_attivi else "Nessun filtro"

        # Definisci le sezioni del report (in questo caso solo una)
        headers = ["Data", "Descrizione", "Conto Finanziario", "Causale", "Entrata", "Uscita", "Anagrafica", "Cantiere"]
        data_rows = []
        for movimento in movimenti_qs:
            data_rows.append([
                movimento.data_registrazione,
                movimento.descrizione,
                movimento.conto_finanziario.nome_conto,
                movimento.causale.descrizione,
                movimento.importo if movimento.tipo_movimento == 'E' else None,
                movimento.importo if movimento.tipo_movimento == 'U' else None,
                str(movimento.anagrafica) if movimento.anagrafica else "",
                str(movimento.cantiere) if movimento.cantiere else "",
            ])
        
        report_sections = [{'title': 'Dettaglio Movimenti', 'headers': headers, 'rows': data_rows}]
        
        # 3. CHIAMATA ALLA FUNZIONE DI UTILITY
        # Deleghiamo tutta la complessità della creazione del file Excel.
        return generate_excel_report(
            tenant_name=tenant_name,
            report_title=report_title,
            filters_string=filtri_str,
            kpi_data=None, # Questo report non ha KPI
            report_sections=report_sections,
            filename_prefix=filename_prefix
        )
  
class PrimaNotaListExportPdfView(PrimaNotaListView):
    """
    Esporta la lista filtrata dei movimenti di Prima Nota in formato PDF.
    """
    def get(self, request, *args, **kwargs):
        # 1. Riutilizziamo la nostra logica centralizzata per ottenere i dati filtrati.
        movimenti_qs, filter_form = self._get_filtered_data(request)
        
        # 2. Prepariamo la stringa dei filtri applicati.
        filtri_attivi = []
        if filter_form.is_valid() and filter_form.cleaned_data:
            for name, value in filter_form.cleaned_data.items():
                if value:
                    label = filter_form.fields[name].label or name.title()
                    # Gestiamo i diversi tipi di campo per avere una visualizzazione pulita
                    if isinstance(value, date):
                        display_value = value.strftime('%d/%m/%Y')
                    elif isinstance(value, models.Model):
                        display_value = str(value)
                    else:
                        display_value = value
                    filtri_attivi.append(f"{label}: {display_value}")
        filtri_str = " | ".join(filtri_attivi) if filtri_attivi else "Nessun filtro"

        # 3. Prepariamo il contesto da passare al template PDF.
        context = {
            'tenant_name': request.session.get('active_tenant_name', 'GestionaleDjango'),
            'report_title': 'Elenco Movimenti di Prima Nota',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            'movimenti': movimenti_qs,
        }
        
        # 4. Chiamiamo la nostra utility per generare il PDF.
        return generate_pdf_report(
            request, 
            'gestionale/primanota_list_pdf.html', 
            context
        )
    
# ==============================================================================
# === VISTE TESORERIA                                                       ===
# ==============================================================================

class TesoreriaDashboardView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Gestisce la visualizzazione della Dashboard di Tesoreria.
    Contiene la logica di calcolo dei saldi riutilizzabile dagli export.
    """
    template_name = 'gestionale/tesoreria_dashboard.html'

    def _get_tesoreria_data(self):
        """
        Metodo helper che calcola i saldi per tutti i conti finanziari.
        """
        conti_finanziari = ContoFinanziario.objects.filter(attivo=True).annotate(
            saldo=Coalesce(
                Sum(
                    models.Case(
                        models.When(movimenti__tipo_movimento='E', then=models.F('movimenti__importo')),
                        models.When(movimenti__tipo_movimento='U', then=-models.F('movimenti__importo')),
                        default=Value(0),
                        output_field=models.DecimalField()
                    )
                ), 
                Value(0),
                output_field=models.DecimalField()
            )
        ).order_by('nome_conto')
        
        liquidita_totale = sum(conto.saldo for conto in conti_finanziari)
        
        return conti_finanziari, liquidita_totale

    def get(self, request, *args, **kwargs):
        conti_finanziari, liquidita_totale = self._get_tesoreria_data()
        
        context = {
            'conti_finanziari': conti_finanziari,
            'liquidita_totale': liquidita_totale,
            'today': date.today(), # Aggiungiamo 'today' che serve nel template
        }
        
        return render(request, self.template_name, context)
    
class TesoreriaExportExcelView(TesoreriaDashboardView):
    """
    Esporta i saldi di tesoreria in formato Excel.
    """
    def get(self, request, *args, **kwargs):
        conti_finanziari, liquidita_totale = self._get_tesoreria_data()
        
        tenant_name = request.session.get('active_tenant_name', 'GestionaleDjango')


        report_title = 'Report Saldi di Tesoreria'
        filename_prefix = 'Saldi_Tesoreria'
        
        kpi_report = {
            'Liquidità Totale': liquidita_totale
        }
        
        headers = ["Conto Finanziario", "Saldo"]
        data_rows = []
        for conto in conti_finanziari:
            data_rows.append([
                conto.nome_conto,
                conto.saldo
            ])
            
        report_sections = [{'title': 'Dettaglio Saldi', 'headers': headers, 'rows': data_rows}]
        
        return generate_excel_report(
            tenant_name, report_title, "Dati al " + timezone.now().strftime('%d/%m/%Y'), 
            kpi_report, report_sections, filename_prefix=filename_prefix
        )

class TesoreriaExportPdfView(TesoreriaDashboardView):
    """
    Esporta i saldi di tesoreria in formato PDF.
    """
    def get(self, request, *args, **kwargs):
        conti_finanziari, liquidita_totale = self._get_tesoreria_data()
        
        context = {
            'tenant_name': request.session.get('active_tenant_name'),
            'report_title': 'Report Saldi di Tesoreria',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': "Dati al " + timezone.now().strftime('%d/%m/%Y'),
            'conti_finanziari': conti_finanziari,
            'liquidita_totale': liquidita_totale,
        }
        
        return generate_pdf_report(
            request, 
            'gestionale/tesoreria_dashboard_pdf.html', 
            context
        )
    

from .templatetags import currency_filters # Importiamo i nostri filtri

class GetContoFinanziarioSaldoView(TenantRequiredMixin, View):
    """
    Vista API che restituisce il saldo di un conto finanziario in formato JSON.
    """
    def get(self, request, *args, **kwargs):
        conto_id = request.GET.get('conto_id')
        if not conto_id:
            return JsonResponse({'error': 'ID Conto mancante'}, status=400)

        try:
            # Riutilizziamo la nostra potente query dalla Tesoreria
            conto = ContoFinanziario.objects.annotate(
                saldo=Coalesce(Sum(Case(When(movimenti__tipo_movimento='E', then=F('movimenti__importo')),When(movimenti__tipo_movimento='U', then=-F('movimenti__importo')),default=Value(0),output_field=models.DecimalField())),Value(0),output_field=models.DecimalField())
            ).get(pk=conto_id)
            
            saldo_formattato = currency_filters.format_currency(conto.saldo)
            
            data = {
                'saldo_formattato': f"Saldo: € {saldo_formattato}",
                'color': 'red' if conto.saldo < 0 else 'green'
            }
            return JsonResponse(data)
        except ContoFinanziario.DoesNotExist:
            return JsonResponse({'error': 'Conto non trovato'}, status=404)

# ==============================================================================
# === VISTE PANNELLO AMMINISTRAZIONE                                        ===
# ==============================================================================

class AdminDashboardView(TenantRequiredMixin, AdminRequiredMixin, View):
    allowed_roles = ['admin']
    """
    Mostra la pagina principale del pannello di amministrazione del tenant.
    """
    template_name = 'gestionale/admin_dashboard.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)
    
# --- VISTE CRUD PER MODALITA' DI PAGAMENTO ---

class ModalitaPagamentoListView(TenantRequiredMixin, AdminRequiredMixin, ListView):
    model = ModalitaPagamento
    template_name = 'gestionale/config_list_base.html' # Useremo un template generico
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modalità di Pagamento'
        context['create_url'] = reverse_lazy('modalita_pagamento_create')
        return context

class ModalitaPagamentoCreateView(TenantRequiredMixin, AdminRequiredMixin, CreateView):
    model = ModalitaPagamento
    form_class = ModalitaPagamentoForm
    template_name = 'gestionale/config_form_base.html' # Template generico
    success_url = reverse_lazy('modalita_pagamento_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuova Modalità di Pagamento'
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Modalità di pagamento creata con successo.")
        return super().form_valid(form)

class ModalitaPagamentoUpdateView(TenantRequiredMixin, AdminRequiredMixin, UpdateView):
    model = ModalitaPagamento
    form_class = ModalitaPagamentoForm
    template_name = 'gestionale/config_form_base.html' # Template generico
    success_url = reverse_lazy('modalita_pagamento_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Modalità di Pagamento'
        return context

    def form_valid(self, form):
        messages.success(self.request, "Modalità di pagamento aggiornata con successo.")
        return super().form_valid(form)

class ModalitaPagamentoToggleAttivoView(TenantRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(ModalitaPagamento, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.descrizione}' aggiornato.")
        return redirect('modalita_pagamento_list')
    
# --- VISTE CRUD PER ALIQUOTE IVA ---

class AliquotaIVAListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = AliquotaIVA
    template_name = 'gestionale/aliquota_iva_list.html' # Useremo un template specifico per ora
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Aliquote IVA'
        context['create_url'] = reverse_lazy('aliquota_iva_create')
        return context

class AliquotaIVACreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = AliquotaIVA
    form_class = AliquotaIVAForm
    template_name = 'gestionale/config_form_base.html' # Riutilizziamo il form generico!
    success_url = reverse_lazy('aliquota_iva_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuova Aliquota IVA'
        # Passiamo l'URL di "annulla" per il template generico
        context['cancel_url'] = reverse_lazy('aliquota_iva_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Aliquota IVA creata con successo.")
        return super().form_valid(form)

class AliquotaIVAUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    model = AliquotaIVA
    form_class = AliquotaIVAForm
    template_name = 'gestionale/config_form_base.html' # Riutilizziamo il form generico!
    success_url = reverse_lazy('aliquota_iva_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Aliquota IVA'
        context['cancel_url'] = reverse_lazy('aliquota_iva_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Aliquota IVA aggiornata con successo.")
        return super().form_valid(form)

class AliquotaIVAToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(AliquotaIVA, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.descrizione}' aggiornato.")
        return redirect('aliquota_iva_list')

# --- VISTE CRUD PER CAUSALI CONTABILI ---

class CausaleListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = Causale
    template_name = 'gestionale/causale_list.html' # Template specifico per la lista
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Causali Contabili'
        context['create_url'] = reverse_lazy('causale_create')
        return context

class CausaleCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = Causale
    form_class = CausaleForm
    template_name = 'gestionale/config_form_base.html' # Riutilizziamo il form generico
    success_url = reverse_lazy('causale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuova Causale Contabile'
        context['cancel_url'] = reverse_lazy('causale_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Causale creata con successo.")
        return super().form_valid(form)

class CausaleUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    model = Causale
    form_class = CausaleForm
    template_name = 'gestionale/config_form_base.html' # Riutilizziamo il form generico
    success_url = reverse_lazy('causale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Causale Contabile'
        context['cancel_url'] = reverse_lazy('causale_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Causale aggiornata con successo.")
        return super().form_valid(form)

class CausaleToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(Causale, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.descrizione}' aggiornato.")
        return redirect('causale_list')

# --- VISTE CRUD PER CONTI FINANZIARI ---

class ContoFinanziarioListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = ContoFinanziario
    template_name = 'gestionale/conto_finanziario_list.html'
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Conti Finanziari (Casse/Banche)'
        context['create_url'] = reverse_lazy('conto_finanziario_create')
        return context

class ContoFinanziarioCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = ContoFinanziario
    form_class = ContoFinanziarioForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('conto_finanziario_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Conto Finanziario'
        context['cancel_url'] = reverse_lazy('conto_finanziario_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Conto Finanziario creato con successo.")
        return super().form_valid(form)

class ContoFinanziarioUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    model = ContoFinanziario
    form_class = ContoFinanziarioForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('conto_finanziario_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Conto Finanziario'
        context['cancel_url'] = reverse_lazy('conto_finanziario_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Conto Finanziario aggiornato con successo.")
        return super().form_valid(form)

class ContoFinanziarioToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(ContoFinanziario, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.nome_conto}' aggiornato.")
        return redirect('conto_finanziario_list')
    
# --- VISTE CRUD PER CONTI OPERATIVI ---

class ContoOperativoListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = ContoOperativo
    template_name = 'gestionale/conto_operativo_list.html'
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Conti Operativi (Costi/Ricavi)'
        context['create_url'] = reverse_lazy('conto_operativo_create')
        return context

class ContoOperativoCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = ContoOperativo
    form_class = ContoOperativoForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('conto_operativo_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Conto Operativo'
        context['cancel_url'] = reverse_lazy('conto_operativo_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Conto Operativo creato con successo.")
        return super().form_valid(form)

class ContoOperativoUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']

    model = ContoOperativo
    form_class = ContoOperativoForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('conto_operativo_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Conto Operativo'
        context['cancel_url'] = reverse_lazy('conto_operativo_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Conto Operativo aggiornato con successo.")
        return super().form_valid(form)

class ContoOperativoToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(ContoOperativo, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.nome_conto}' aggiornato.")
        return redirect('conto_operativo_list')
    
 # --- VISTE CRUD PER MEZZI AZIENDALI ---

class MezzoAziendaleListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = MezzoAziendale
    template_name = 'gestionale/mezzo_aziendale_list.html'
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Gestione Flotta Mezzi'
        context['create_url'] = reverse_lazy('mezzo_aziendale_create')
        return context

class MezzoAziendaleCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = MezzoAziendale
    form_class = MezzoAziendaleForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('mezzo_aziendale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Mezzo Aziendale'
        context['cancel_url'] = reverse_lazy('mezzo_aziendale_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Mezzo aziendale creato con successo.")
        return super().form_valid(form)

class MezzoAziendaleUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    model = MezzoAziendale
    form_class = MezzoAziendaleForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('mezzo_aziendale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Mezzo Aziendale'
        context['cancel_url'] = reverse_lazy('mezzo_aziendale_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Mezzo aziendale aggiornato con successo.")
        return super().form_valid(form)

class MezzoAziendaleToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(MezzoAziendale, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.targa}' aggiornato.")
        return redirect('mezzo_aziendale_list')
   
# --- VISTE CRUD PER TIPI SCADENZE PERSONALE ---

class TipoScadenzaPersonaleListView(TenantRequiredMixin, RoleRequiredMixin, ListView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    model = TipoScadenzaPersonale
    template_name = 'gestionale/tipo_scadenza_personale_list.html'
    context_object_name = 'oggetti'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Tipi Scadenze del Personale'
        context['create_url'] = reverse_lazy('tipo_scadenza_personale_create')
        return context

class TipoScadenzaPersonaleCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin']
    model = TipoScadenzaPersonale
    form_class = TipoScadenzaPersonaleForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('tipo_scadenza_personale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Tipo Scadenza Personale'
        context['cancel_url'] = reverse_lazy('tipo_scadenza_personale_list')
        return context
        
    def form_valid(self, form):
        messages.success(self.request, "Tipo scadenza creato con successo.")
        return super().form_valid(form)

class TipoScadenzaPersonaleUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    model = TipoScadenzaPersonale
    form_class = TipoScadenzaPersonaleForm
    template_name = 'gestionale/config_form_base.html'
    success_url = reverse_lazy('tipo_scadenza_personale_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Modifica Tipo Scadenza Personale'
        context['cancel_url'] = reverse_lazy('tipo_scadenza_personale_list')
        return context

    def form_valid(self, form):
        messages.success(self.request, "Tipo scadenza aggiornato con successo.")
        return super().form_valid(form)

class TipoScadenzaPersonaleToggleAttivoView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    def post(self, request, *args, **kwargs):
        obj = get_object_or_404(TipoScadenzaPersonale, pk=kwargs.get('pk'))
        obj.attivo = not obj.attivo
        obj.save()
        messages.success(request, f"Stato di '{obj.descrizione}' aggiornato.")
        return redirect('tipo_scadenza_personale_list')
    
class DipendenteDetailView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore'] # Cambia da DetailView a View
    """
    Mostra il "Fascicolo del Dipendente", la pagina di dettaglio completa
    per un'anagrafica di tipo Dipendente, con tabelle paginate.
    """
    template_name = 'gestionale/dipendente_detail.html'

    def get(self, request, *args, **kwargs):
        """
        Gestisce la richiesta GET per la pagina del Fascicolo Dipendente.
        """
        # 1. RECUPERO DELL'OGGETTO PRINCIPALE
        # Usiamo get_object_or_404 per recuperare il dipendente in modo sicuro,
        # assicurandoci che esista e che sia effettivamente di tipo 'Dipendente'.
        dipendente = get_object_or_404(Anagrafica, pk=kwargs['pk'], tipo=Anagrafica.Tipo.DIPENDENTE)
        dettaglio = dipendente.dettaglio_dipendente
        ore_default = 0
        if dettaglio and dettaglio.giorni_lavorativi_settimana > 0:
            ore_default = round(dettaglio.ore_settimanali_contratto / dettaglio.giorni_lavorativi_settimana, 1)

        # 2. RECUPERO DEI DATI COLLEGATI
        # Recupera lo storico completo delle attività, pre-caricando i dati
        # del cantiere per ottimizzare le query.
        storico_attivita_qs = DiarioAttivita.objects.filter(
            dipendente=dipendente
        ).select_related('cantiere_pianificato').order_by('-data')
        
        # Recupera le scadenze personali, pre-caricando i dati del tipo di scadenza.
        scadenze_personali_qs = ScadenzaPersonale.objects.filter(
            dipendente=dipendente
        ).select_related('tipo_scadenza').order_by('data_scadenza')

        # 3. PREPARAZIONE DATI PER LOGICA JAVASCRIPT
        # Creiamo un dizionario che mappa {id_tipo_scadenza: validita_mesi}
        # per la logica di auto-calcolo della data di scadenza.
        # Includiamo solo i tipi che hanno una validità in mesi definita e > 0.
        tipi_scadenza_attivi = TipoScadenzaPersonale.objects.filter(attivo=True)
        tipi_scadenza_data = {
            tipo.id: tipo.validita_mesi 
            for tipo in tipi_scadenza_attivi 
            if tipo.validita_mesi is not None and tipo.validita_mesi > 0
        }

        # 4. GESTIONE DELLA PAGINAZIONE INDIPENDENTE
        # Paginazione per lo Storico Attività
        paginator_attivita = Paginator(storico_attivita_qs, 10)
        page_number_attivita = request.GET.get('pagina_attivita', 1)
        page_obj_attivita = paginator_attivita.get_page(page_number_attivita)

        # Paginazione per le Scadenze Personali
        paginator_scadenze = Paginator(scadenze_personali_qs, 5)
        page_number_scadenze = request.GET.get('pagina_scadenze', 1)
        page_obj_scadenze = paginator_scadenze.get_page(page_number_scadenze)

        # 5. PREPARAZIONE DEL CONTESTO FINALE
        context = {
            'title': f"Fascicolo Dipendente: {dipendente.nome_cognome_ragione_sociale}",
            'dipendente': dipendente,
            'storico_attivita': page_obj_attivita,
            'scadenze_personali': page_obj_scadenze,
            'today': date.today(), # Utile per evidenziare le scadenze scadute
            'scadenza_personale_form': ScadenzaPersonaleForm(), # Form per la modale di creazione
            'tipi_scadenza_data_json': json.dumps(tipi_scadenza_data), # Dati per lo script JS
            'attivita_form': DiarioAttivitaForm(),
            'ore_giornaliere_default': ore_default,
        }
        
        # 6. RENDERIZZAZIONE DEL TEMPLATE
        return render(request, self.template_name, context)
    
# ==============================================================================
# === VISTE CRUD SCADENZE PERSONALI (per il Fascicolo Dipendente)           ===
# ==============================================================================

class ScadenzaPersonaleCreateView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin']
    """
    Gestisce la creazione di una nuova scadenza personale.
    """
    def post(self, request, *args, **kwargs):
        dipendente = get_object_or_404(Anagrafica, pk=kwargs['dipendente_pk'])
        form = ScadenzaPersonaleForm(request.POST)
        if form.is_valid():
            scadenza = form.save(commit=False)
            scadenza.dipendente = dipendente
            scadenza.created_by = request.user
            scadenza.save()
            messages.success(request, "Scadenza personale creata con successo.")
        else:
            messages.error(request, f"Errore nella creazione della scadenza: {form.errors.as_text()}")
        return redirect('dipendente_detail', pk=dipendente.pk)

class ScadenzaPersonaleUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin']
    """
    Gestisce la modifica di una scadenza personale esistente.
    """
    model = ScadenzaPersonale
    form_class = ScadenzaPersonaleForm
    template_name = 'gestionale/scadenza_personale_form.html'
    
    def get_success_url(self):
        return reverse_lazy('dipendente_detail', kwargs={'pk': self.object.dipendente.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Modifica Scadenza per {self.object.dipendente.nome_cognome_ragione_sociale}"
        
        tipi_scadenza_attivi = TipoScadenzaPersonale.objects.filter(attivo=True)
        tipi_scadenza_data = {
            tipo.id: tipo.validita_mesi 
            for tipo in tipi_scadenza_attivi
            if tipo.validita_mesi is not None and tipo.validita_mesi > 0
        }
        context['tipi_scadenza_data_json'] = json.dumps(tipi_scadenza_data)
        
        return context

    def form_valid(self, form):
        messages.success(self.request, "Scadenza personale aggiornata con successo.")
        return super().form_valid(form)

class ScadenzaPersonaleDeleteView(TenantRequiredMixin, RoleRequiredMixin, DeleteView):
    allowed_roles = ['admin']
    """
    Gestisce l'eliminazione di una scadenza personale.
    """
    model = ScadenzaPersonale
    template_name = 'gestionale/generic_confirm_delete.html' # Un template di conferma generico
    
    def get_success_url(self):
        messages.success(self.request, "Scadenza personale eliminata con successo.")
        return reverse_lazy('dipendente_detail', kwargs={'pk': self.object.dipendente.pk})
    
# ==============================================================================
# === DASHBOARD PRINCIPALE                                                   ===
# ==============================================================================

class DashboardView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Mostra la dashboard principale con KPI aggregati e riepiloghi operativi.
    """
    template_name = 'gestionale/dashboard.html'

    def get(self, request, *args, **kwargs):
        today = date.today()
        current_year = today.year
        
        # --- 1. CALCOLI FINANZIARI ESISTENTI (Scadenze e Saldi) ---
        scadenze_aperte_qs = Scadenza.objects.filter(
            stato__in=[Scadenza.Stato.APERTA, Scadenza.Stato.PARZIALE]
        )

        kpi_rate = scadenze_aperte_qs.aggregate(
            incassi_tot_rate=Coalesce(Sum('importo_rata', filter=Q(tipo_scadenza='Incasso')), Value(0), output_field=DecimalField()),
            pagamenti_tot_rate=Coalesce(Sum('importo_rata', filter=Q(tipo_scadenza='Pagamento')), Value(0), output_field=DecimalField()),
        )
        
        kpi_pagamenti = PrimaNota.objects.filter(scadenza_collegata__in=scadenze_aperte_qs).aggregate(
            incassi_pagati=Coalesce(Sum('importo', filter=Q(scadenza_collegata__tipo_scadenza='Incasso')), Value(0), output_field=DecimalField()),
            pagamenti_pagati=Coalesce(Sum('importo', filter=Q(scadenza_collegata__tipo_scadenza='Pagamento')), Value(0), output_field=DecimalField()),
        )

        incassi_aperti = kpi_rate['incassi_tot_rate'] - kpi_pagamenti['incassi_pagati']
        debiti_aperti = kpi_rate['pagamenti_tot_rate'] - kpi_pagamenti['pagamenti_pagati']

        scaduti_agg = scadenze_aperte_qs.filter(data_scadenza__lt=today).annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Decimal('0.00'))
        ).aggregate(
            crediti_scaduti=Coalesce(Sum(F('importo_rata') - F('pagato'), filter=Q(tipo_scadenza='Incasso')), Decimal('0.00')),
            debiti_scaduti=Coalesce(Sum(F('importo_rata') - F('pagato'), filter=Q(tipo_scadenza='Pagamento')), Decimal('0.00')),
        )
        
        conti_finanziari = ContoFinanziario.objects.filter(attivo=True).annotate(
            saldo=Coalesce(Sum(Case(When(movimenti__tipo_movimento='E', then=F('movimenti__importo')), When(movimenti__tipo_movimento='U', then=-F('movimenti__importo')), default=Value(0), output_field=DecimalField())), Value(0), output_field=DecimalField())
        ).order_by('nome_conto')
        
        liquidita_totale = sum(c.saldo for c in conti_finanziari)

        # --- 2. NUOVI KPI: ECONOMICI E ANAGRAFICI ---

        kpi_anagrafiche = Anagrafica.objects.aggregate(
            clienti=Count('id', filter=Q(tipo=Anagrafica.Tipo.CLIENTE, attivo=True)),
            fornitori=Count('id', filter=Q(tipo=Anagrafica.Tipo.FORNITORE, attivo=True)),
            dipendenti=Count('id', filter=Q(tipo=Anagrafica.Tipo.DIPENDENTE, attivo=True))
        )

        fatturato_ytd_agg = DocumentoTestata.objects.filter(
            data_documento__year=current_year,
            stato=DocumentoTestata.Stato.CONFERMATO
        ).aggregate(
            fatture_vendita=Coalesce(Sum('totale', filter=Q(tipo_doc=DocumentoTestata.TipoDoc.FATTURA_VENDITA)), Decimal('0.00')),
            note_credito_vendita=Coalesce(Sum('totale', filter=Q(tipo_doc=DocumentoTestata.TipoDoc.NOTA_CREDITO_VENDITA)), Decimal('0.00')),
            fatture_acquisto=Coalesce(Sum('totale', filter=Q(tipo_doc=DocumentoTestata.TipoDoc.FATTURA_ACQUISTO)), Decimal('0.00')),
            note_credito_acquisto=Coalesce(Sum('totale', filter=Q(tipo_doc=DocumentoTestata.TipoDoc.NOTA_CREDITO_ACQUISTO)), Decimal('0.00'))
        )
        fatturato_attivo = fatturato_ytd_agg['fatture_vendita'] - fatturato_ytd_agg['note_credito_vendita']
        fatturato_passivo = fatturato_ytd_agg['fatture_acquisto'] - fatturato_ytd_agg['note_credito_acquisto']
        fatturato_netto_ytd = fatturato_attivo - fatturato_passivo

        cash_flow_ytd_agg = PrimaNota.objects.filter(
            data_registrazione__year=current_year, 
            conto_finanziario__isnull=False
        ).aggregate(
            entrate=Coalesce(Sum('importo', filter=Q(tipo_movimento='E')), Decimal('0.00')),
            uscite=Coalesce(Sum('importo', filter=Q(tipo_movimento='U')), Decimal('0.00'))
        )
        cash_flow_ytd = cash_flow_ytd_agg['entrate'] - cash_flow_ytd_agg['uscite']

        movimenti_economici_ytd = PrimaNota.objects.filter(
            data_registrazione__year=current_year,
            conto_operativo__isnull=False
        ).aggregate(
            ricavi=Coalesce(Sum('importo', filter=Q(conto_operativo__tipo=ContoOperativo.Tipo.RICAVO)), Decimal('0.00')),
            costi=Coalesce(Sum('importo', filter=Q(conto_operativo__tipo=ContoOperativo.Tipo.COSTO)), Decimal('0.00'))
        )
        risultato_economico_ytd = movimenti_economici_ytd['ricavi'] - movimenti_economici_ytd['costi']

        # --- 3. NUOVI KPI: OPERATIVI (Margine per Cantiere) ---
        cantieri_con_margine = Cantiere.objects.filter(stato=Cantiere.Stato.APERTO).annotate(
            ricavi_totali=Coalesce(Sum('movimenti_primanota__importo', filter=Q(movimenti_primanota__conto_operativo__tipo=ContoOperativo.Tipo.RICAVO)), Decimal('0.00')),
            costi_totali=Coalesce(Sum('movimenti_primanota__importo', filter=Q(movimenti_primanota__conto_operativo__tipo=ContoOperativo.Tipo.COSTO)), Decimal('0.00'))
        ).annotate(
            margine=F('ricavi_totali') - F('costi_totali')
        ).order_by('-margine')

        # --- 4. WIDGET ESISTENTI (Scadenze e Riepilogo HR) ---
        sessanta_giorni_da_oggi = today + timedelta(days=60)
        scadenze_imminenti = scadenze_aperte_qs.filter(
            data_scadenza__gte=today,
            data_scadenza__lte=sessanta_giorni_da_oggi
        ).annotate(
            pagato=Coalesce(Sum('pagamenti__importo'), Value(0), output_field=DecimalField()),
            residuo=F('importo_rata') - F('pagato')
        ).filter(residuo__gt=0).order_by('data_scadenza')[:5]

        dipendenti_presenti = DiarioAttivita.objects.filter(data=today, stato_presenza=DiarioAttivita.StatoPresenza.PRESENTE).count()
        totale_dipendenti_attivi = kpi_anagrafiche['dipendenti']
        cantieri_attivi = Cantiere.objects.filter(stato=Cantiere.Stato.APERTO).count()
        
        # --- 5. COSTRUZIONE DEL CONTESTO PER IL TEMPLATE ---
        context = {
            'kpi_finanziari': {
                'crediti_clienti': incassi_aperti,
                'crediti_scaduti': scaduti_agg['crediti_scaduti'],
                'debiti_fornitori': debiti_aperti,
                'debiti_scaduti': scaduti_agg['debiti_scaduti'],
                'liquidita_totale': liquidita_totale,
                'saldo_netto': incassi_aperti - debiti_aperti,
            },
            'kpi_economici_e_anagrafici': {
                'anagrafiche': kpi_anagrafiche,
                'fatturato_netto_ytd': fatturato_netto_ytd,
                'cash_flow_ytd': cash_flow_ytd,
                'risultato_economico_ytd': risultato_economico_ytd,
            },
            'cantieri_con_margine': cantieri_con_margine,
            'conti_finanziari': conti_finanziari,
            'scadenze_imminenti': scadenze_imminenti,
            'kpi_operativi': {
                'dipendenti_presenti': dipendenti_presenti,
                'totale_dipendenti': totale_dipendenti_attivi,
                'cantieri_attivi': cantieri_attivi,
                'note_spese_pendenti': 0, # Placeholder
            }
        }
        
        return render(request, self.template_name, context)

class DashboardAnalisiView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    template_name = 'gestionale/dashboard_analisi.html'

    def _get_kpi_data(self, data_da, data_a):
        """
        Metodo helper che calcola tutti i KPI in base a un intervallo di date.
        Versione COMPLETA con tutti i KPI.
        """
        # ... Sezioni 1 e 2 con i calcoli di stato e di flusso rimangono invariate ...
        # (copio qui il codice corretto per completezza)
        movimenti_fino_a_data_a = PrimaNota.objects.filter(data_registrazione__lte=data_a, conto_finanziario__isnull=False)
        liquidita_totale = movimenti_fino_a_data_a.aggregate(saldo=Coalesce(Sum(Case(When(tipo_movimento='E', then=F('importo')), default=-F('importo'))), Value(0), output_field=DecimalField()))['saldo']
        scadenze_fino_a_data_a = Scadenza.objects.filter(documento__data_documento__lte=data_a, documento__stato=DocumentoTestata.Stato.CONFERMATO)
        pagamenti_fino_a_data_a = PrimaNota.objects.filter(data_registrazione__lte=data_a, scadenza_collegata__isnull=False)
        crediti_emessi = scadenze_fino_a_data_a.filter(tipo_scadenza='Incasso').aggregate(t=Coalesce(Sum('importo_rata'), Value(0), output_field=DecimalField()))['t']
        incassi_ricevuti = pagamenti_fino_a_data_a.filter(scadenza_collegata__tipo_scadenza='Incasso').aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        crediti_clienti = crediti_emessi - incassi_ricevuti
        debiti_ricevuti = scadenze_fino_a_data_a.filter(tipo_scadenza='Pagamento').aggregate(t=Coalesce(Sum('importo_rata'), Value(0), output_field=DecimalField()))['t']
        pagamenti_effettuati = pagamenti_fino_a_data_a.filter(scadenza_collegata__tipo_scadenza='Pagamento').aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        debiti_fornitori = debiti_ricevuti - pagamenti_effettuati
        doc_periodo = DocumentoTestata.objects.filter(data_documento__range=(data_da, data_a), stato=DocumentoTestata.Stato.CONFERMATO)
        fatturato_attivo_periodo = (doc_periodo.filter(tipo_doc__in=['FTV']).aggregate(t=Coalesce(Sum('totale'), Value(0), output_field=DecimalField()))['t']) - (doc_periodo.filter(tipo_doc__in=['NCV']).aggregate(t=Coalesce(Sum('totale'), Value(0), output_field=DecimalField()))['t'])
        costi_fatturati_periodo = (doc_periodo.filter(tipo_doc__in=['FTA']).aggregate(t=Coalesce(Sum('totale'), Value(0), output_field=DecimalField()))['t']) - (doc_periodo.filter(tipo_doc__in=['NCA']).aggregate(t=Coalesce(Sum('totale'), Value(0), output_field=DecimalField()))['t'])
        mov_eco_periodo = PrimaNota.objects.filter(data_registrazione__range=(data_da, data_a), conto_operativo__isnull=False)
        ricavi_periodo = mov_eco_periodo.filter(conto_operativo__tipo=ContoOperativo.Tipo.RICAVO).aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        costi_periodo = mov_eco_periodo.filter(conto_operativo__tipo=ContoOperativo.Tipo.COSTO).aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        risultato_economico_periodo = ricavi_periodo - costi_periodo
        mov_fin_periodo = PrimaNota.objects.filter(data_registrazione__range=(data_da, data_a), conto_finanziario__isnull=False)
        entrate_periodo = mov_fin_periodo.filter(tipo_movimento='E').aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        uscite_periodo = mov_fin_periodo.filter(tipo_movimento='U').aggregate(t=Coalesce(Sum('importo'), Value(0), output_field=DecimalField()))['t']
        cash_flow_periodo = entrate_periodo - uscite_periodo
        
        # === INIZIO MODIFICA: Aggiungiamo il conteggio anagrafiche ===
        anagrafiche_attive = Anagrafica.objects.filter(attivo=True).aggregate(
            clienti=Count('id', filter=Q(tipo=Anagrafica.Tipo.CLIENTE)),
            fornitori=Count('id', filter=Q(tipo=Anagrafica.Tipo.FORNITORE)),
            dipendenti=Count('id', filter=Q(tipo=Anagrafica.Tipo.DIPENDENTE))
        )
        num_dipendenti = anagrafiche_attive['dipendenti']
        # === FINE MODIFICA ===

        # --- 3. KPI DERIVATI (con aggiornamento num_dipendenti) ---
        giorni_periodo = (data_a - data_da).days + 1
        giorni_periodo_decimal = Decimal(giorni_periodo)
        dso = (crediti_clienti / fatturato_attivo_periodo * giorni_periodo_decimal) if fatturato_attivo_periodo > 0 else Decimal('0')
        dpo = (debiti_fornitori / costi_fatturati_periodo * giorni_periodo_decimal) if costi_fatturati_periodo > 0 else Decimal('0')
        mesi_periodo_decimal = giorni_periodo_decimal / Decimal('30.44')
        burn_rate_mensile = (costi_periodo / mesi_periodo_decimal) if mesi_periodo_decimal > 0 else Decimal('0')
        runway_mesi = (liquidita_totale / burn_rate_mensile) if burn_rate_mensile > 0 else float('inf')
        
        ricavi_per_dipendente = (fatturato_attivo_periodo / num_dipendenti) if num_dipendenti > 0 else Decimal('0')
        margine_lordo_perc = (risultato_economico_periodo / ricavi_periodo * 100) if ricavi_periodo > 0 else Decimal('0')

        return {
            'liquidita_totale': liquidita_totale, 'posizione_fin_netta': crediti_clienti - debiti_fornitori,
            'fatturato_netto_periodo': fatturato_attivo_periodo - costi_fatturati_periodo,
            'risultato_economico_periodo': risultato_economico_periodo, 'cash_flow_periodo': cash_flow_periodo,
            'dso': dso, 'dpo': dpo, 'burn_rate_mensile': burn_rate_mensile, 'runway_mesi': runway_mesi,
            'ricavi_per_dipendente': ricavi_per_dipendente, 'margine_lordo_perc': margine_lordo_perc,
            'anagrafiche_attive': anagrafiche_attive, # <-- Aggiungiamo al dizionario
            'crediti_clienti': crediti_clienti,      # <-- Aggiungiamo anche questi per le card
            'debiti_fornitori': debiti_fornitori,    # <-- Aggiungiamo anche questi per le card
        }
    


    def get(self, request, *args, **kwargs):
        filter_form = AnalisiFilterForm(request.GET or None)
        
        # Gestione date di default: Anno Corrente fino a Oggi
        today = date.today()
        data_da_default = date(today.year, 1, 1)
        data_a_default = today

        data_da = data_a = None
        if filter_form.is_valid():
            data_da = filter_form.cleaned_data.get('data_da')
            data_a = filter_form.cleaned_data.get('data_a')

        # Se le date non sono fornite, usiamo i default
        if not data_da: data_da = data_da_default
        if not data_a: data_a = data_a_default

        context = {
            'kpi': self._get_kpi_data(data_da, data_a),
            'filter_form': filter_form,
            'data_da_corrente': data_da,
            'data_a_corrente': data_a,
        }
        return render(request, self.template_name, context)

# ==============================================================================
# === EXPORT TABELLE DI SISTEMA                                              ===
# ==============================================================================
class ExportTabelleSistemaView(TenantRequiredMixin, AdminRequiredMixin, View):
    """
    Gestisce l'esportazione di tutte le tabelle dell'app 'gestionale'
    in un unico file Excel, con un foglio per ogni tabella.
    """
    def get(self, request, *args, **kwargs):
        # 1. PREPARAZIONE DELLA RISPOSTA E DEL FILE EXCEL
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        local_time = timezone.localtime(timezone.now())
        timestamp = local_time.strftime('%Y%m%d%H%M')
        filename = f"{timestamp}_Esportazione_Tabelle_di_Sistema_GestiLub.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook = Workbook()
        workbook.remove(workbook.active) # Rimuoviamo il foglio di default
        header_font = Font(bold=True)

        # 2. IDENTIFICAZIONE DEI MODELLI DA ESPORTARE
        # Usiamo il registro delle app di Django per trovare dinamicamente
        # tutti i modelli definiti nella nostra app 'gestionale'.
        app_models = apps.get_app_config('gestionale').get_models()

        # 3. CICLO DI ESPORTAZIONE PER OGNI MODELLO
        for model in app_models:
            model_name = model.__name__
            queryset = model.objects.all()
            
            # Se non ci sono dati, non creiamo il foglio
            if not queryset.exists():
                continue

            # Crea un nuovo foglio di lavoro per questo modello
            worksheet = workbook.create_sheet(title=model_name[:31]) # Titolo foglio max 31 chars

            # Scrivi le intestazioni (i nomi dei campi del modello)
            headers = [field.name for field in model._meta.fields]
            worksheet.append(headers)
            for cell in worksheet[1]:
                cell.font = header_font

            # Scrivi le righe di dati
            for obj in queryset:
                row_data = []
                for field_name in headers:
                    cell_value = getattr(obj, field_name)

                    # Controlla se il valore è un oggetto datetime
                    if isinstance(cell_value, datetime):
                        # Se è timezone-aware, convertilo nel fuso orario locale
                        if timezone.is_aware(cell_value):
                            cell_value = timezone.localtime(cell_value)
                        # Rendi l'oggetto datetime "naive" (senza tzinfo)
                        cell_value = cell_value.replace(tzinfo=None)
                    
                    # Gestisci i campi ForeignKey
                    elif isinstance(cell_value, models.Model):
                        cell_value = str(cell_value)

                    row_data.append(cell_value)
                worksheet.append(row_data)

        # Se non abbiamo esportato nessun foglio, creane uno vuoto di avviso
        if not workbook.sheetnames:
            worksheet = workbook.create_sheet(title="NessunDato")
            worksheet.append(["Nessun dato trovato in nessuna tabella dell'applicazione."])

        # 4. SALVATAGGIO E RESTITUZIONE
        workbook.save(response)
        return response

class ExportTabelleContabiliView(TenantRequiredMixin, AdminRequiredMixin, View):
    """
    Gestisce l'esportazione di tutte le tabelle operative/contabili
    dell'azienda corrente in un unico file Excel.
    """
    def get(self, request, *args, **kwargs):
        # 1. PREPARAZIONE DELLA RISPOSTA E DEL FILE EXCEL
        tenant_name = request.session.get('active_tenant_name', 'Gestionale')
        safe_tenant_name = "".join(c for c in tenant_name if c.isalnum() or c in " _-").rstrip()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        local_time = timezone.localtime(timezone.now())
        timestamp = local_time.strftime('%Y%m%d%H%M')
        filename = f"{timestamp}_Esportazione_Tabelle_Contabili_{safe_tenant_name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook = Workbook()
        workbook.remove(workbook.active)
        header_font = Font(bold=True)

        # 2. DEFINIZIONE ESPLICITA DEI MODELLI DA ESPORTARE
        # A differenza dell'export di sistema, qui selezioniamo a mano i modelli.
        models_to_export = [
            Anagrafica, Cantiere, DocumentoTestata, DocumentoRiga, Scadenza,
            PrimaNota, DipendenteDettaglio, DiarioAttivita, ScadenzaPersonale
        ]

        # 3. CICLO DI ESPORTAZIONE (identico alla vista precedente)
        for model in models_to_export:
            model_name = model.__name__
            queryset = model.objects.all()
            
            if not queryset.exists():
                continue

            worksheet = workbook.create_sheet(title=model_name[:31])
            headers = [field.name for field in model._meta.fields]
            worksheet.append(headers)
            for cell in worksheet[1]:
                cell.font = header_font

            for obj in queryset:
                row_data = []
                for field_name in headers:
                    cell_value = getattr(obj, field_name)

                    if isinstance(cell_value, datetime):
                        if timezone.is_aware(cell_value):
                            cell_value = timezone.localtime(cell_value)
                        cell_value = cell_value.replace(tzinfo=None)
                    
                    elif isinstance(cell_value, models.Model):
                        cell_value = str(cell_value)

                    row_data.append(cell_value)
                worksheet.append(row_data)

        if not workbook.sheetnames:
            worksheet = workbook.create_sheet(title="NessunDato")
            worksheet.append(["Nessun dato trovato nelle tabelle contabili."])

        # 4. SALVATAGGIO E RESTITUZIONE
        workbook.save(response)
        return response
    
class DipendenteUpdateView(TenantRequiredMixin, AdminRequiredMixin, View):
    """
    Gestisce la modifica di un dipendente, che coinvolge due modelli:
    Anagrafica e DipendenteDettaglio.
    """
    template_name = 'gestionale/dipendente_update_form.html'

    def get(self, request, *args, **kwargs):
        # Recuperiamo l'anagrafica del dipendente
        dipendente = get_object_or_404(Anagrafica, pk=kwargs['pk'], tipo=Anagrafica.Tipo.DIPENDENTE)
        
        # Creiamo le istanze dei due form, popolandole con i dati esistenti
        anagrafica_form = AnagraficaForm(instance=dipendente)
        dettaglio_form = DipendenteDettaglioForm(instance=dipendente.dettaglio_dipendente)

        context = {
            'title': f"Modifica Dipendente: {dipendente.nome_cognome_ragione_sociale}",
            'dipendente': dipendente,
            'anagrafica_form': anagrafica_form,
            'dettaglio_form': dettaglio_form
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        dipendente = get_object_or_404(Anagrafica, pk=kwargs['pk'], tipo=Anagrafica.Tipo.DIPENDENTE)

        # Creiamo le istanze dei form con i dati POST e le istanze originali
        anagrafica_form = AnagraficaForm(request.POST, instance=dipendente)
        dettaglio_form = DipendenteDettaglioForm(request.POST, instance=dipendente.dettaglio_dipendente)

        # Validiamo entrambi i form
        if anagrafica_form.is_valid() and dettaglio_form.is_valid():
            with transaction.atomic():
                # Salviamo il primo form
                anagrafica_aggiornata = anagrafica_form.save(commit=False)
                anagrafica_aggiornata.updated_by = request.user
                anagrafica_aggiornata.save()

                # Salviamo il secondo form
                dettaglio_aggiornato = dettaglio_form.save(commit=False)
                # Il metodo save() del dettaglio si occuperà di aggiornare lo stato 'attivo'
                dettaglio_aggiornato.save()

            messages.success(request, "Dati del dipendente aggiornati con successo.")
            return redirect('dipendente_detail', pk=dipendente.pk)
        else:
            # Se uno dei due form non è valido, ripresentiamo la pagina con gli errori
            context = {
                'title': f"Modifica Dipendente: {dipendente.nome_cognome_ragione_sociale}",
                'dipendente': dipendente,
                'anagrafica_form': anagrafica_form,
                'dettaglio_form': dettaglio_form
            }
            messages.error(request, "Correggi gli errori nel form.")
            return render(request, self.template_name, context)
        
# ==============================================================================
# === VISTE CRUD CANTIERI                                                   ===
# ==============================================================================

class CantiereCreateView(TenantRequiredMixin, RoleRequiredMixin, CreateView):
    allowed_roles = ['admin', 'contabile']
    model = Cantiere
    form_class = CantiereForm
    template_name = 'gestionale/cantiere_form.html' # Useremo un template dedicato
    
    def get_success_url(self):
        # Torniamo alla dashboard HR da cui siamo partiti
        return reverse_lazy('dashboard_hr')

    def form_valid(self, form):
        cantiere = form.save(commit=False)
        cantiere.created_by = self.request.user
        cantiere.updated_by = self.request.user
        cantiere.save()
        messages.success(self.request, "Cantiere creato con successo.")
        return HttpResponseRedirect(self.get_success_url())
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuovo Cantiere'
        return context

class CantiereUpdateView(TenantRequiredMixin, RoleRequiredMixin, UpdateView):
    allowed_roles = ['admin', 'contabile']
    model = Cantiere
    form_class = CantiereForm
    template_name = 'gestionale/cantiere_form.html'
    
    def get_success_url(self):
        return reverse_lazy('dashboard_hr')

    def form_valid(self, form):
        cantiere = form.save(commit=False)
        cantiere.updated_by = self.request.user
        cantiere.save()
        messages.success(self.request, "Cantiere aggiornato con successo.")
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Modifica Cantiere: {self.object.codice_cantiere}"
        return context
    
class CantiereListExportExcelView(TenantRequiredMixin, AdminRequiredMixin, View):
    """
    Esporta la lista filtrata dei cantieri in formato Excel.
    """
    def get(self, request, *args, **kwargs):
        # Riutilizziamo la logica di filtraggio dalla DashboardHRView
        stato_cantiere_filter = request.GET.get('stato_cantiere', Cantiere.Stato.APERTO)
        cantieri_qs = Cantiere.objects.select_related('cliente').order_by('codice_cantiere')
        if stato_cantiere_filter and stato_cantiere_filter != 'TUTTI':
            cantieri_qs = cantieri_qs.filter(stato=stato_cantiere_filter)
            
        # Preparazione dati per l'utility
        tenant_name = request.session.get('active_tenant_name')
        report_title = 'Elenco Cantieri'
        filename_prefix = 'Elenco_Cantieri'
        
        filtri_str = f"Stato: {dict(Cantiere.Stato.choices).get(stato_cantiere_filter, 'Tutti')}"
        
        headers = ["Codice", "Descrizione", "Cliente", "Indirizzo", "Stato", "Data Inizio", "Data Fine Prevista"]
        data_rows = []
        for cantiere in cantieri_qs:
            data_rows.append([
                cantiere.codice_cantiere,
                cantiere.descrizione,
                cantiere.cliente.nome_cognome_ragione_sociale,
                cantiere.indirizzo,
                cantiere.get_stato_display(),
                cantiere.data_inizio,
                cantiere.data_fine_prevista
            ])
            
        report_sections = [{'title': 'Dettaglio Cantieri', 'headers': headers, 'rows': data_rows}]
        
        return generate_excel_report(
            tenant_name, report_title, filtri_str, None, report_sections,
            filename_prefix=filename_prefix
        )

class CantiereListExportPdfView(TenantRequiredMixin, AdminRequiredMixin, View):
    """
    Esporta la lista filtrata dei cantieri in formato PDF.
    """
    def get(self, request, *args, **kwargs):
        # Riutilizziamo la logica di filtraggio
        stato_cantiere_filter = request.GET.get('stato_cantiere', Cantiere.Stato.APERTO)
        cantieri_qs = Cantiere.objects.select_related('cliente').order_by('codice_cantiere')
        if stato_cantiere_filter and stato_cantiere_filter != 'TUTTI':
            cantieri_qs = cantieri_qs.filter(stato=stato_cantiere_filter)

        filtri_str = f"Stato: {dict(Cantiere.Stato.choices).get(stato_cantiere_filter, 'Tutti')}"

        context = {
            'tenant_name': request.session.get('active_tenant_name'),
            'report_title': 'Elenco Cantieri',
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            'cantieri': cantieri_qs,
        }
        
        return generate_pdf_report(request, 'gestionale/cantiere_list_pdf.html', context)
    
class SuperAdminRequiredMixin(AccessMixin):
    """
    Mixin per verificare che l'utente sia loggato e abbia il ruolo
    di sistema 'super_admin'.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.system_role == 'super_admin':
            messages.error(request, "Accesso negato. Area riservata ai Super Amministratori.")
            # Se un utente normale prova ad accedere, lo mandiamo alla selezione tenant.
            return redirect(reverse('tenant_selection')) 
        return super().dispatch(request, *args, **kwargs)
    

# ==============================================================================
# === VISTE FASCICOLO CANTIERE (DETAIL + EXPORTS)                          ===
# ==============================================================================

class CantiereDetailView(TenantRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Gestisce la visualizzazione del Fascicolo Cantiere e contiene la logica
    di recupero dati riutilizzata dagli export.
    """
    template_name = 'gestionale/cantiere_detail.html'

    def _get_fascicolo_data(self, request, cantiere_pk):
        """
        Metodo helper che recupera e filtra TUTTI i dati per il fascicolo.
        """
        cantiere = get_object_or_404(Cantiere, pk=cantiere_pk)
        filter_form = FascicoloCantiereFilterForm(request.GET or None)

        # Queryset di base
        dipendenti_qs = DiarioAttivita.objects.filter(cantiere_pianificato=cantiere).select_related('dipendente').order_by('-data')
        documenti_qs = DocumentoTestata.objects.filter(cantiere=cantiere, stato=DocumentoTestata.Stato.CONFERMATO).order_by('-data_documento')
        movimenti_qs = PrimaNota.objects.filter(cantiere=cantiere).select_related('conto_finanziario').order_by('-data_registrazione')

        if filter_form.is_valid():
            data_da = filter_form.cleaned_data.get('data_da')
            data_a = filter_form.cleaned_data.get('data_a')
            if data_da:
                dipendenti_qs = dipendenti_qs.filter(data__gte=data_da)
                documenti_qs = documenti_qs.filter(data_documento__gte=data_da)
                movimenti_qs = movimenti_qs.filter(data_registrazione__gte=data_da)
            if data_a:
                dipendenti_qs = dipendenti_qs.filter(data__lte=data_a)
                documenti_qs = documenti_qs.filter(data_documento__lte=data_a)
                movimenti_qs = movimenti_qs.filter(data_registrazione__lte=data_a)
        
        # === NUOVA SEZIONE: CALCOLO RIASSUNTIVO ===
    
        # 1. Calcolo Fatturato e Costi da Documenti
        documenti_cantiere = DocumentoTestata.objects.filter(cantiere=cantiere, stato=DocumentoTestata.Stato.CONFERMATO)
        
        tot_fatture_vendita = documenti_cantiere.filter(tipo_doc=DocumentoTestata.TipoDoc.FATTURA_VENDITA).aggregate(tot=Sum('totale'))['tot'] or 0
        tot_nc_vendita = documenti_cantiere.filter(tipo_doc=DocumentoTestata.TipoDoc.NOTA_CREDITO_VENDITA).aggregate(tot=Sum('totale'))['tot'] or 0
        fatturato_netto = tot_fatture_vendita - tot_nc_vendita

        tot_fatture_acquisto = documenti_cantiere.filter(tipo_doc=DocumentoTestata.TipoDoc.FATTURA_ACQUISTO).aggregate(tot=Sum('totale'))['tot'] or 0
        tot_nc_acquisto = documenti_cantiere.filter(tipo_doc=DocumentoTestata.TipoDoc.NOTA_CREDITO_ACQUISTO).aggregate(tot=Sum('totale'))['tot'] or 0
        costi_fatturati_netti = tot_fatture_acquisto - tot_nc_acquisto

        # 2. Calcolo Incassi e Pagamenti da Prima Nota (legati a scadenze del cantiere)
        movimenti_cantiere = PrimaNota.objects.filter(cantiere=cantiere)
        
        incassi_totali = movimenti_cantiere.filter(tipo_movimento=PrimaNota.TipoMovimento.ENTRATA).aggregate(tot=Sum('importo'))['tot'] or 0
        pagamenti_totali = movimenti_cantiere.filter(tipo_movimento=PrimaNota.TipoMovimento.USCITA).aggregate(tot=Sum('importo'))['tot'] or 0

        # 3. Calcolo Costi/Ricavi diretti da Prima Nota (se non derivanti da fatture)
        costi_diretti = movimenti_cantiere.filter(
            conto_operativo__tipo=ContoOperativo.Tipo.COSTO
        ).aggregate(tot=Sum('importo'))['tot'] or 0
        
        ricavi_diretti = movimenti_cantiere.filter(
            conto_operativo__tipo=ContoOperativo.Tipo.RICAVO
        ).aggregate(tot=Sum('importo'))['tot'] or 0

        # 4. KPI
        redditivita = (fatturato_netto + ricavi_diretti) - (costi_fatturati_netti + costi_diretti)
        cash_flow = incassi_totali - pagamenti_totali
        esposizione_clienti = fatturato_netto - incassi_totali
        esposizione_fornitori = costi_fatturati_netti - pagamenti_totali

        riepilogo = {
            "fatturato_netto": fatturato_netto,
            "costi_fatturati_netti": costi_fatturati_netti,
            "incassi_totali": incassi_totali,
            "pagamenti_totali": pagamenti_totali,
            "costi_diretti": costi_diretti,
            "ricavi_diretti": ricavi_diretti,
            "redditivita": redditivita,
            "cash_flow": cash_flow,
            "esposizione_clienti": esposizione_clienti,
            "esposizione_fornitori": esposizione_fornitori,
        }
        # === FINE NUOVA SEZIONE ===

        return {
            "cantiere": cantiere,
            "filter_form": filter_form,
            "dipendenti_assegnati": dipendenti_qs,
            "documenti_associati": documenti_qs,
            "movimenti_associati": movimenti_qs,
            "riepilogo": riepilogo, # Aggiungiamo il riepilogo al dizionario
        }

    def get(self, request, *args, **kwargs):
        """Metodo per la visualizzazione della pagina HTML."""
        fascicolo_data = self._get_fascicolo_data(request, kwargs['pk'])
        context = {}
        context.update(fascicolo_data)

        # Paginazione per i dipendenti
        paginator_dip = Paginator(fascicolo_data['dipendenti_assegnati'], 10)
        page_dip = request.GET.get('pagina_dipendenti', 1)
        context['dipendenti_assegnati'] = paginator_dip.get_page(page_dip)

        # Paginazione per i documenti
        paginator_doc = Paginator(fascicolo_data['documenti_associati'], 10)
        page_doc = request.GET.get('pagina_documenti', 1)
        context['documenti_associati'] = paginator_doc.get_page(page_doc)

        # Paginazione per i movimenti
        paginator_mov = Paginator(fascicolo_data['movimenti_associati'], 10)
        page_mov = request.GET.get('pagina_movimenti', 1)
        context['movimenti_associati'] = paginator_mov.get_page(page_mov)

        return render(request, self.template_name, context)
class CantiereFascicoloExportExcelView(CantiereDetailView):
    """
    Esporta il fascicolo cantiere filtrato in formato Excel.
    Eredita da CantiereDetailView per riutilizzare _get_fascicolo_data.
    """
    allowed_roles = ['admin', 'contabile', 'visualizzatore'] # Assicurati che questa riga sia presente
    
    def get(self, request, *args, **kwargs):
        fascicolo_data = self._get_fascicolo_data(request, kwargs['pk'])
        cantiere = fascicolo_data['cantiere']
        riepilogo = fascicolo_data['riepilogo'] # <-- Recuperiamo il dizionario con i KPI

        tenant_name = request.session.get('active_tenant_name', 'N/A')
        report_title = f"Fascicolo Cantiere: {cantiere.codice_cantiere}"
        safe_cantiere_name = "".join(c for c in cantiere.codice_cantiere if c.isalnum()).rstrip()
        filename_prefix = f"Fascicolo_{safe_cantiere_name}"
        filtri_str = build_filters_string(fascicolo_data['filter_form'])

        # === INIZIO MODIFICA: PREPARAZIONE DEI KPI PER L'EXPORT ===
        # Creiamo un dizionario ordinato e con etichette leggibili per il report
        kpi_report = {
            "Redditività (Ricavi - Costi)": riepilogo['redditivita'],
            "Cash Flow (Incassi - Pagamenti)": riepilogo['cash_flow'],
            "Crediti vs Clienti": riepilogo['esposizione_clienti'],
            "Debiti vs Fornitori": riepilogo['esposizione_fornitori']
        }
        # === FINE MODIFICA ===

        report_sections = []

        # Sezione 1: Dipendenti (invariata)
        dip_headers = ["Data", "Dipendente", "Stato", "Ore Ordinarie", "Ore Straordinarie"]
        dip_rows = [[d.data, d.dipendente.nome_cognome_ragione_sociale, d.get_stato_presenza_display(), d.ore_ordinarie, d.ore_straordinarie] for d in fascicolo_data['dipendenti_assegnati']]
        report_sections.append({'title': 'Personale Assegnato', 'headers': dip_headers, 'rows': dip_rows})

        # Sezione 2: Documenti (invariata)
        doc_headers = ["Data", "Tipo", "Numero", "Anagrafica", "Totale"]
        doc_rows = [[d.data_documento, d.get_tipo_doc_display(), d.numero_documento, d.anagrafica.nome_cognome_ragione_sociale, d.totale] for d in fascicolo_data['documenti_associati']]
        report_sections.append({'title': 'Documenti Associati', 'headers': doc_headers, 'rows': doc_rows})

        # Sezione 3: Movimenti (invariata)
        mov_headers = ["Data", "Descrizione", "Importo", "Conto", "Causale"]
        mov_rows = [[m.data_registrazione, m.descrizione, m.importo * (1 if m.tipo_movimento == 'E' else -1), m.conto_finanziario.nome_conto, m.causale.descrizione] for m in fascicolo_data['movimenti_associati']]
        report_sections.append({'title': 'Movimenti di Prima Nota Associati', 'headers': mov_headers, 'rows': mov_rows})

        # MODIFICA: Passiamo kpi_report alla funzione generate_excel_report
        return generate_excel_report(tenant_name, report_title, filtri_str, kpi_report, report_sections, filename_prefix)
    

class CantiereFascicoloExportPdfView(CantiereDetailView):
    allowed_roles = ['admin', 'contabile', 'visualizzatore']
    """
    Esporta il fascicolo cantiere filtrato in formato PDF.
    """
    def get(self, request, *args, **kwargs):
        fascicolo_data = self._get_fascicolo_data(request, kwargs['pk'])
        cantiere = fascicolo_data['cantiere']
        filtri_str = build_filters_string(fascicolo_data['filter_form'])
        
        context = {
            'tenant_name': request.session.get('active_tenant_name', 'N/A'),
            'report_title': f"Fascicolo Cantiere",
            'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'filtri_str': filtri_str,
            **fascicolo_data
        }
        
        return generate_pdf_report(request, 'gestionale/cantiere_fascicolo_pdf.html', context)