# gestionale/report_utils.py

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML

# ==============================================================================
# === UTILITY PER EXPORT EXCEL                                              ===
# ==============================================================================

# gestionale/report_utils.py

def generate_excel_report(tenant_name, report_title, filters_string, kpi_data, report_sections, filename_prefix='report'):
    """
    Funzione generica DEFINITIVA per creare un report Excel strutturato.
    Usa un contatore di riga manuale per un controllo totale sul layout e la spaziatura.
    """
    # 1. PREPARAZIONE (invariato)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    local_time = timezone.localtime(timezone.now())
    timestamp = local_time.strftime('%Y%m%d-%H%M')
    filename = f"{timestamp}-{filename_prefix}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = Workbook()
    worksheet = workbook.active
    safe_sheet_title = report_title.replace(":", "-").replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
    worksheet.title = safe_sheet_title[:31]
    
    # STILI (invariato)
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True)
    section_title_font = Font(name='Calibri', size=13, bold=True, italic=True)
    currency_format = '€ #,##0.00'
    
    # ==========================================================================
    # === INIZIO LOGICA DI SCRITTURA ESPLICITA                               ===
    # ==========================================================================
    
    # Usiamo un contatore di riga manuale. Partiamo dalla riga 1.
    current_row = 1
    max_cols = max(len(s.get('headers', [1])) for s in report_sections) if report_sections else 1

    # 2. SCRITTURA INTESTAZIONE
    worksheet.cell(row=current_row, column=1, value=tenant_name).font = title_font
    worksheet.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
    current_row += 1
    
    worksheet.cell(row=current_row, column=1, value=report_title).alignment = Alignment(horizontal='center')
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
    current_row += 2 # Riga vuota

    worksheet.cell(row=current_row, column=1, value=f"Filtri Applicati: {filters_string}")
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
    current_row += 1
    
    worksheet.cell(row=current_row, column=1, value=f"Generato il: {local_time.strftime('%d/%m/%Y %H:%M:%S')}")
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
    current_row += 1

    # 3. SCRITTURA KPI
    if kpi_data:
        current_row += 1 # Riga vuota prima dei KPI
        for key, value in kpi_data.items():
            worksheet.cell(row=current_row, column=2, value=key).font = header_font
            cell = worksheet.cell(row=current_row, column=3, value=value)
            cell.number_format = currency_format
            current_row += 1

    # 4. RIGA VUOTA DOPO LA TESTATA
    current_row += 1
    
    # 5. SCRITTURA DELLE SEZIONI
    for section in report_sections:
        # 5.1 TITOLO SEZIONE
        worksheet.cell(row=current_row, column=1, value=section.get('title', '')).font = section_title_font
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
        current_row += 1
        
        # 5.2 RIGA VUOTA DOPO IL TITOLO
        current_row += 1
        
        # 5.3 INTESTAZIONI DELLA TABELLA
        headers = section.get('headers', [])
        if headers:
            for col_idx, header_title in enumerate(headers, 1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header_title)
                cell.font = header_font
            current_row += 1
            
        # 5.4 RIGHE DI DATI
        for row_data in section.get('rows', []):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=current_row, column=col_idx, value=cell_value)
            current_row += 1
            
        # 5.5 RIGA VUOTA DOPO LA FINE DELLE RIGHE DELLA SEZIONE
        current_row += 1

    # 6. ADATTAMENTO COLONNE (invariato)
    for col_idx in range(1, max_cols + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in worksheet[column_letter]:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value or ""))
            except: pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
        
    # 7. SALVATAGGIO E RESTITUZIONE
    workbook.save(response)
    return response



# ==============================================================================
# === UTILITY PER EXPORT PDF                                              ===
# ==============================================================================

def generate_pdf_report(request, template_name, context, filename=None):
    """
    Funzione generica per creare un report PDF da un template HTML.

    Args:
        request: L'oggetto richiesta di Django.
        template_name (str): Il percorso del template HTML da usare.
        context (dict): Il dizionario di contesto con i dati.
        filename (str, optional): Il nome file desiderato per il download.
                                  Se non fornito, ne viene generato uno di default.

    Returns:
        HttpResponse: La risposta HTTP con il file PDF da scaricare.
    """
    response = HttpResponse(content_type='application/pdf')
    
    # Se un nome file non è fornito, ne crea uno di default.
    if not filename:
        report_title = context.get('report_title', 'report')
        timestamp = timezone.now().strftime('%Y%m%d')
        filename = f"{report_title.replace(' ', '_').lower()}_{timestamp}.pdf"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    html_string = render_to_string(template_name, context, request=request)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response.write(pdf_file)
    return response


